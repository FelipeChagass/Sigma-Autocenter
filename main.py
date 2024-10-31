from PySide6.QtWidgets import *
from PySide6.QtGui import QIcon, QPainter, QKeyEvent
from PySide6.QtCore import *
from PySide6 import QtCore, QtWidgets
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtSql import QSqlDatabase, QSqlTableModel
import functools
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import brazilcep
from PIL.ImageQt import ImageQt
import sys
import sqlite3 
import os
import aspose.pdf as pdf
from datetime import date
import re 
import random
import tempfile
from pdf2image import convert_from_path
import matplotlib.pyplot as plt
from Login import Ui_Login
from Layout_Sigma_Admin import Ui_MainWindow
from Layout_Sigma_Usuario import Ui_MainWindowUsuario
from database import Database
from ler import Read_xml
import Imagens
import Icones
import xml.etree.ElementTree as ET

class Login(QMainWindow, Ui_Login):
    def __init__(self):
        super(Login, self).__init__()
        self.setupUi(self)
        appIcon = QIcon('Imagens/Estatua.png')
        self.setWindowIcon(appIcon)
        self.showFullScreen()
        self.tentativas = 0
        db = Database()
        db.connect()
        db.create_table_funcionario()
        db.create_table_trabalhos()
        db.create_table_nfe()
        db.create_table_cliente()
        db.create_table_fornecedor()
        db.insert_admin()
        self.setWindowTitle("Login do Sistema") 
        self.btn_entrar.clicked.connect(self.checklogin)
        #Se clicar no X do Popup, ele fecha
        self.pushButton_close_popup.clicked.connect(lambda: self.frame_erro.hide())
        #Se clicar no botão, ele chama a função de checagem de campos
        self.btn_entrar.clicked.connect(self.checkFields)
        #O frame se esconde hihi
        self.frame_erro.hide()
    
    #
    #Estilos
    StyleLineEditOk = ("QLineEdit{\n"
"    border: 2px solid rgb(45,45,45);\n"
"    border-radius: 5px;\n"
"    padding: 10px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")

    StyleLineEditErro = ("QLineEdit{\n"
"    border: 2px solid rgb(255, 21, 21);\n"
"    border-radius: 5px;\n"
"    padding: 10px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")
    
    #Style Popup
    StylePopupErro = ("background-color: rgb(208, 9, 9); border-radius: 58x;")
    

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_F11:
            self.showFullScreen() if not self.isFullScreen() else self.showNormal()

    #
    #Funções da Tela de Login
    def checkFields(self):
    
        def showMessage(message):
            self.frame_erro.show()
            self.label_erro.setText(message)

        #checka Usuario
        if not self.txt_usuario_login.text():
            TextoPopupUsuario = "Campo usuário está vazio."
            self.txt_usuario_login.setStyleSheet(self.StyleLineEditErro)
        else:
            TextoPopupUsuario = ""
            self.txt_usuario_login.setStyleSheet(self.StyleLineEditOk)
        #Checka Senha
        if not self.txt_senha_login.text():
                 TextoPopupSenha = " Campo senha está vazio."
                 self.txt_senha_login.setStyleSheet(self.StyleLineEditErro)
        else:
                TextoPopupSenha= ""
                self.txt_senha_login.setStyleSheet(self.StyleLineEditOk)

        #Checka os campos
        if TextoPopupUsuario + TextoPopupSenha != '':
             text = TextoPopupUsuario + TextoPopupSenha 
             showMessage(text)
             self.frame_erro.setStyleSheet(self.StylePopupErro) 

    def checklogin(self):
        self.users = Database()
        self.users.connect()
        autenticado = self.users.check_user(self.txt_usuario_login.text(), self.txt_senha_login.text())

        if autenticado.lower() == "administrador":
            self.w = TelaADM()
            self.w.show()
            self.close()
        elif autenticado == "Usuário":
            self.w = MainWindowUsuario()
            self.w.show()
            self.close() 
        else:
            if self.tentativas <3:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Erro Ao Acessar")
                msg.setText(f'Login ou senha incorreto \n \n Tentativa: {self.tentativas +1} de 3')
                msg.exec()
                self.tentativas += 1
                if self.tentativas == 3:
                    self.users.close_connection()
                    sys.exit(0)

class TelaADM(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(TelaADM, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Sistema de Gerenciamento")
        self.showFullScreen()
        appIcon = QIcon('Imagens/Estatua.png')
        self.setWindowIcon(appIcon)

        #Muda as Páginas do Sistema
        self.btn_menu.clicked.connect(lambda: self.moverMenuLateral())
        self.btn_menu_cadastro_usuario.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.Page_CadastroUsuario))
        self.btn_menu_cadastro_usuario.clicked.connect(lambda:self.Reset_Funcionario())
        self.btn_menu_fornecedores.clicked.connect(lambda: self.Reset_Fornecedores ())
        self.btn_menu_nota_fiscal.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.Page_Importar_XML))
        self.btn_menu_config.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_ajustes))
        self.btn_confUser.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_usuarios))
        self.btn_adicionar_cliente.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_clientes))
        self.btn_menu_clientes.clicked.connect(lambda: self.Reset_Clientes())
        self.btn_menu_clientes.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_visualizar_clientes))
        self.btn_menu_config.clicked.connect(lambda:self.ApagarLabelErros())
        self.btn_adicionar_item.clicked.connect(lambda: self.Reset_Produtos())
        self.btn_menu_estoque.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_Estoque))
        self.btn_menu_fornecedores.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_fornecedor))
        self.btn_adicionar_fornecedor.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_adicionar_fornecedor))
        self.btn_adicionar_fornecedor.clicked.connect(lambda: self.Reset_Fornecedores())
        self.btn_sair.clicked.connect(lambda: self.Sair())
        self.btn_menu_trabalho.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_trabalhos))
        self.btn_adicionar_trabalho.clicked.connect(lambda: self.Reset_Trabalhos())
        self.btn_adicionar_trabalho.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_AdicionarTrabalho))

        #Botões Cadastro Cliente
        self.btn_cancelar_cadastro.clicked.connect(lambda: self.Cancelar())
        self.btn_salvar_cadastro.clicked.connect(lambda: self.cadastrar_usuario()) 
        self.btn_olho_senha.clicked.connect(lambda:self.toggleVisibility())
        self.btn_olho_senha_confirmar_senha.clicked.connect(lambda:self.toggleVisibility2())

        #Botões Adicionar Trabalho
        self.btn_cancelar_trabalho.clicked.connect(lambda: self.CancelarTrabalho())
        self.btn_salvar_trabalho.clicked.connect(lambda: self.Adicionar_Trabalho())

        #Botões Adicionar Produtos
        self.btn_salvar_produto.clicked.connect(lambda: self.Adicionar_Item())
        self.btn_cancelar_produto.clicked.connect(lambda: self.CancelarProduto())

        #Botões Adicionar Clientes
        self.btn_salvar_cliente.clicked.connect(lambda: self.Adicionar_Cliente())

        # Botões Adicionar Fornecedores
        self.txt_cep_fornecedor.returnPressed.connect(self.CepFornecedor)
        self.btn_salvar_fornecedor.clicked.connect(self.Adicionar_fornecedor)
        self.btn_adicionar_fornecedor.clicked.connect(self.Reset_Fornecedores())

        # Instala um filtro de eventos no campo de texto
        self.txt_cep_fornecedor.installEventFilter(self)

        #Esconder Botões
        self.btn_excluir.hide()

        #Arquivo XML
        self.btn_abrir_diretorio.clicked.connect(self.open_path)
        self.btn_importar_xml.clicked.connect(self.import_arquivos_xml)

        #tabelas
        self.btn_gerar_saida.clicked.connect(self.gerar_saida)
        self.btn_gerar_estorno.clicked.connect(self.gerar_estorno)
        self.btn_realizar_trabalho.clicked.connect(self.realizar_trabalho)
        self.btn_voltar_trabalho.clicked.connect(self.voltar_trabalho)
        self.reset_table()

        #Botôes Estoque
        self.btn_adicionar_item.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.Page_Produto))
        self.btn_gerar_excel.clicked.connect(lambda: self.Converter_Excel())
        self.btn_grafico.clicked.connect(lambda: self.graphic())
        self.btn_imprimir.clicked.connect(lambda: self.onPrint())
        self.btn_excluir.clicked.connect(lambda: self.questionExcluir())


        # ComboBoxes
        self.cmb_pecas.currentTextChanged.connect(self.atualizar_id_funcionario)
        self.cmb_cliente.currentTextChanged.connect(self.atualizar_cpf_cliente)
        self.cmb_funcionario_responsavel.currentTextChanged.connect(self.atualizar_id_funcionario)
        self.Chave_Nota = None
        self.btn_adicionar_cliente.clicked.connect(self.carregar_produtos_na_combobox)
        self.btn_adicionar_trabalho.clicked.connect(self.carregar_usuario_na_combobox_funcionario)
        self.btn_adicionar_trabalho.clicked.connect(self.carregar_nome_na_combobox_cliente)

        #Updates
        self.updating_cliente = False
        self.updating_funcionario = False
        self.updating_fornecedor = False
        self.updating_trabalho = False

        #Filtros
        self.setup_connections_cliente_filtro()  #configura a filtragem
        self.setup_connections_funcionario_filtro()
        self.setup_connections_trabalhos_filtro()
        self.setup_connections_fornecedor_filtro()
        self.txt_filtro.textChanged.connect(self.update_filter)


        #Cancelar
        self.btn_cancelar_cliente.clicked.connect(self.Cancelar_Cliente)
        self.btn_cancelar_fornecedor.clicked.connect(self.Cancelar_Fornecedores)
        self.btn_cancelar_trabalho.clicked.connect(self.Cancelar_Trabalho)


    #Estilos
    StyleLineEditOk = ("QLineEdit{\n"
"    border: 2px solid rgb(47, 82, 162);\n"
"    border-radius: 5px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")

    StyleLineEditErro = ("QLineEdit{\n"
"    border: 2px solid rgb(255, 21, 21);\n"
"    border-radius: 5px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")


    def Sair(self):
                   
            self.w = Login()
            self.w.show()
            self.close()


    def ApagarLabelErros(self):
                   
        self.lbl_erro_endereco_vazio.hide()
        self.lbl_erro_nome_bd_vazio.hide()
        self.lbl_erro_usuario_bd_vazio.hide()
        self.lbl_erro_senha_bd_vazio.hide()

   
    def Reset_Funcionario(self):
        self.lbl_titulo_cadastro_usuario.setText("Cadastro de Funcionarios")
        self.btn_salvar_cadastro.setText("Salvar")
        self.txt_nome_usuario.setText("")
        self.txt_usuario.setText("")
        self.txt_senha_usuario.setText("")
        self.txt_confirmar_senha.setText("")
        self.txt_nome_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_senha_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_vazio.hide()
        self.lbl_erro_usuario_vazio.hide()
        self.lbl_erro_senha_vazio.hide()
        self.lbl_erro_confirmar_senha_vazio.hide()

    def Reset_Clientes(self):
        self.lbl_titulo_cadastro_cliente.setText("Cadastro de Clientes")
        self.btn_salvar_cliente.setText("Salvar")
        self.txt_cpf_cliente.setText("")
        self.txt_nome_cliente.setText("")
        self.txt_telefone_cliente.setText("")
        self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditOk)
        self.txt_nome_cliente.setStyleSheet(self.StyleLineEditOk)
        self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_cliente_vazio.hide()
        self.lbl_erro_cpf_vazio.hide()
        self.lbl_erro_telefone_vazio.hide()
        self.lbl_erro_telefone_vazio.hide()

    def Reset_Trabalhos(self):
        self.lbl_titulo_cadastro_trabalho.setText("Cadastro de Trabalhos")
        self.btn_salvar_trabalho.setText("Salvar")
        self.txt_nome_trabalho.setText("")
        self.txt_modelo_auto.setText("")
        self.txt_preco_servico.setText("")
        self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditOk)
        self.txt_modelo_auto.setStyleSheet(self.StyleLineEditOk)
        self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_trabalho_vazio.hide()
        self.lbl_erro_modeloauto_vazio.hide()
        self.lbl_erro_preco_vazio.hide()

    def Reset_Fornecedores(self):
        self.lbl_titulo_fornecedor.setText("Fornecedores")
        self.btn_salvar_fornecedor.setText("Salvar")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_IE_fornecedor.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_email_fornecedor.setText("")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_telefone_fornecedor.setText("")
        self.txt_cep_fornecedor.setText("")
        self.txt_numero_fornecedor.setText("")
        self.txt_logradouro_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")
        self.txt_UF_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")    
        self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)
        self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)
        self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)
        self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)
        self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_cep_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_fantasia_vazio.hide()
        self.lbl_erro_razao_social_vazio.hide()
        self.lbl_erro_cnpj_vazio.hide()
        self.lbl_erro_IE_vazio.hide()
        self.lbl_erro_telefone_fornecedor_vazio.hide()
        self.lbl_erro_email_fornecedor_vazio.hide()
        self.lbl_erro_numero_casa_vazio.hide()
        self.lbl_erro_logradouro_vazio.hide()
        self.lbl_erro_bairro_vazio.hide()
        self.lbl_erro_cidade_vazio.hide()

    def Reset_Produtos(self):
        self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_unidade_medida.setStyleSheet(self.StyleLineEditOk)
        self.txt_descricao_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_codigo_produto.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_qtd_produto_vazio.hide()
        self.lbl_erro_codigo_produto_vazio.hide()
        self.lbl_erro_descricao_produto_vazio.hide()
        self.lbl_erro_unidade_medida_vazio.hide()
        self.lbl_erro_preco_produto_vazio.hide()


    def Cancelar_Cliente(self):
        self.txt_cpf_cliente.setText("")
        self.txt_nome_cliente.setText("")
        self.txt_telefone_cliente.setText("")

    def Cancelar_Fornecedores(self):
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_IE_fornecedor.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_email_fornecedor.setText("")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_telefone_fornecedor.setText("")
        self.txt_cep_fornecedor.setText("")
        self.txt_numero_fornecedor.setText("")
        self.txt_logradouro_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")
        self.txt_UF_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")    

    def Cancelar_Trabalho(self):
        self.txt_nome_trabalho.setText("")
        self.txt_modelo_auto.setText("")
        self.txt_preco_servico.setText("")


    def keyPressEvent(self, event: QKeyEvent):
        
        if event.key() == Qt.Key_F11: 
            if self.isFullScreen(): 
                self.showNormal() 
            else: self.showFullScreen() 
        elif event.key() == Qt.Key_Escape and self.isFullScreen(): 
            self.showNormal()


    def moverMenuLateral(self):
         width = self.Left_Menu.width()

         # Se estiver minimizado
         if width == 80:
              #Expandir Menu
              newWidth = 250
         else:
              #VoltaAoNormal
              newWidth = 80

        #Animar Transição
         self.animation = QtCore.QPropertyAnimation (self.Left_Menu, b"maximumWidth")
         self.animation.setDuration(250)
         self.animation.setStartValue(width)
         self.animation.setEndValue(newWidth)
         self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
         self.animation.start()


    def toggleVisibility(self):
        if self.txt_senha_usuario.echoMode()==QLineEdit.Normal:
            self.txt_senha_usuario.setEchoMode(QLineEdit.Password)
        else:
            self.txt_senha_usuario.setEchoMode(QLineEdit.Normal)
            
    def toggleVisibility2(self):
        if self.txt_confirmar_senha.echoMode()==QLineEdit.Normal:
            self.txt_confirmar_senha.setEchoMode(QLineEdit.Password)
        else:
            self.txt_confirmar_senha.setEchoMode(QLineEdit.Normal)


    def cadastrar_usuario(self):
         
        #Checka se as senhas são iguais
        if self.txt_senha_usuario.text() != self.txt_confirmar_senha.text():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Senhas Conflitantes")
            msg.setText("Senhas Divergentes")
            msg.exec_()
            return None
         
        if not self.txt_nome_usuario.text():
            self.txt_nome_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_usuario.text():
            self.txt_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_usuario_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_senha_usuario.text():
            self.txt_senha_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_senha_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_senha_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_confirmar_senha.text():
            self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_confirmar_senha_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditOk)              

       
        Nome = self.txt_nome_usuario.text()
        Usuario = self.txt_usuario.text()
        Senha = self.txt_senha_usuario.text()
        Acesso = self.cmb_perfil.currentText()

        db = Database()
        db.connect()


        if not getattr(self, 'updating_funcionario', False):
            # Verifica se o funcionario já existe no banco de dados apenas se não estiver atualizando
            funcionario_existe = db.check_if_user_exists(Usuario)

            if funcionario_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Funcionário Existente")
                msg.setText(f"O funcionário '{Usuario}' já existe.")
                msg.exec_()
                self.txt_cpf_cliente.setText("")
                self.txt_nome_cliente.setText("")
                self.txt_telefone_cliente.setText("")
                self.lbl_erro_nome_cliente_vazio.hide()
                self.lbl_erro_cpf_vazio.hide()
                self.lbl_erro_telefone_vazio.hide()

        # Marca a variável _updating_funcionario como False após a atualização
        if getattr(self, 'updating_funcionario', False):
            setattr(self, 'updating_funcionario', False)
            ID = self.tw_funcionario.selectedItems()[0].text(0) 
            self.update_funcionario(ID, Nome, Usuario, Senha, Acesso)

        else:
            length_ok = len(Senha) >= 8
            uppercase_ok = any(c.isupper() for c in Senha)
            lowercase_ok = any(c.islower() for c in Senha)
            digit_ok = any(c.isdigit() for c in Senha)
            special_ok = any(not c.isalnum() for c in Senha)
            error_messages = []

            if not length_ok:
                error_messages.append("A senha requer 8 caracteres")
            if not uppercase_ok:
                error_messages.append("A senha requer letras maiúsculas")
            if not lowercase_ok:
                error_messages.append("A senha requer letras minúsculas")
            if not digit_ok:
                error_messages.append("A senha requer pelo menos um número")
            if not special_ok:
                error_messages.append("A senha requer pelo menos um caractere especial")

            if error_messages:
                strength = "Fraca" # Ou outra força adequada, dependendo dos erros
                self.lbl_erro_senha_vazio.show()
                error_message = f"Resultado da verificação: Senha {strength}\n" + "\n".join(error_messages)
                self.lbl_erro_senha_vazio.setText(error_message)
                self.txt_senha_usuario.setStyleSheet(self.StyleLineEditErro)
            else:
                # Se não houver erros, você consegue cadastrar o usuário
                strength = "Muito Forte"
                db.insert_funcionario(Nome, Usuario, Senha, Acesso)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Cadastro de Funcionários")
                msg.setText("Funcionário Adicionado com Sucesso")
                msg.exec_()
                self.txt_nome_usuario.setText("")
                self.txt_usuario.setText("")
                self.txt_senha_usuario.setText("")
                self.txt_confirmar_senha.setText("")
                self.lbl_erro_nome_vazio.hide()
                self.lbl_erro_usuario_vazio.hide()
                self.lbl_erro_senha_vazio.hide()
                self.lbl_erro_confirmar_senha_vazio.hide()
                self.table_funcionario()


    def table_funcionario(self):

        self.tw_funcionario.clear()

        connection = sqlite3.connect("AutoCenter.db")

        cursor = connection.cursor()

        # Pega as informações da tabela funcionário
        cursor.execute("SELECT * FROM funcionario")
        records = cursor.fetchall()

        for record in records:
            item = QTreeWidgetItem(self.tw_funcionario)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_funcionario(item, record))
            self.tw_funcionario.setItemWidget(item, 5, delete_button)  # Ajuste na posição

            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.edit_selected_funcionario())
            edit_button.clicked.connect(lambda: self.ApagarLabelErros())
            self.tw_funcionario.setItemWidget(item, 6, edit_button)  # Ajuste na posição

        # Resize nas colunas
        self.tw_funcionario.resizeColumnToContents(0)  
        for i in range(1, self.tw_cliente.columnCount()):  
            self.tw_funcionario.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_funcionario.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        connection.close()


    def setup_connections_funcionario_filtro(self):
        self.txt_filtro_funcionario.textChanged.connect(self.filter_table_funcionario)

    def filter_table_funcionario(self, text):
        for row in range(self.tw_funcionario.topLevelItemCount()):
            item = self.tw_funcionario.topLevelItem(row)
            nome_index_funcionario = 1 #Coluna onde está o nome do funcionário

            if item.text(nome_index_funcionario).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def edit_selected_funcionario(self):
        selected_items = self.tw_funcionario.selectedItems()  # Obter itens selecionados na tabela
        
        # Perguntar ao usuário se eles desejam editar o funcionario selecionado
        reply = QMessageBox.question(self, "Editar Funcionário", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o ID do funcionario selecionado
            ID = selected_items[0].text(0) 

            # Obter os detalhes do Funcionario no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM funcionario WHERE ID = ?", (ID,))
            funcionario = cursor.fetchone()
            connection.close()

            # Definir a página atual como a página de funcionários
            self.Paginas.setCurrentWidget(self.Page_CadastroUsuario)

            # Definir o texto dos campos com os detalhes do funcionário
            self.txt_nome_usuario.setText(funcionario[1])
            self.txt_usuario.setText(funcionario[2])
            self.txt_senha_usuario.setText(funcionario[3])
            self.txt_confirmar_senha.setText(funcionario[3])
            self.btn_salvar_cadastro.setText("Editar")
            self.lbl_titulo_cadastro_usuario.setText("Editar Funcionario")


            if self.cmb_perfil.count() > 0:
                index = self.cmb_perfil.findText(funcionario[4])
                if index != -1:
                    self.cmb_perfil.setCurrentIndex(index)

            self.updating_funcionario = True


    def update_funcionario(self, ID, Nome, Usuario, Senha, Acesso):

        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE funcionario
            SET Nome=?, Usuario=?, Senha=?, Acesso=?
            WHERE ID=?
        """, (Nome, Usuario, Senha, Acesso, ID))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Edição de Funcionário")
        msg.setText("Funcionário Editado com Sucesso")
        msg.exec_()

        # Clear input fields
        self.txt_nome_usuario.clear()
        self.txt_usuario.clear()
        self.txt_senha_usuario.clear()
        self.txt_confirmar_senha.clear()
        self.table_funcionario()

    def on_btn_salvar_funcionario_clicked(self):
        
        Nome = self.txt_nome_usuario.text()
        Usuario = self.txt_usuario.text()
        Senha = self.txt_senha_usuario.text()
        Acesso = self.cmb_perfil.currentText()

        # Checa se o funcionário está selecionado
        selected_items = self.tw_funcionario.selectedItems()
        print("Selected Items:", selected_items)

        if selected_items:
            # Se o funcionario estiver selecionado, edita
            ID = selected_items[0].text(0)

            db = Database()
            db.connect()
            db.close_connection()

            self.update_funcionario()
        else:
            # Se nenhum funcionario estiver selecionado, salva normalmente
            db.insert_funcionario(Nome, Usuario, Senha, Acesso)

    def excluir_funcionario(self, ID):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Funcionário", "Deseja realmente excluir o funcionário?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o funcionario do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM funcionario WHERE ID=?", (ID,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de funcionários após a exclusão
            self.table_funcionario()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Funcionário")
            msg.setText("Funcionário Excluído Com Sucesso")
            msg.exec_()

    def confirmar_delete_funcionario(self, item, record):

        # Obter o ID do funcionario a ser excluído
        ID = record[0]

        # Chamar a função para excluir o funcionario
        self.excluir_funcionario(ID)


    def Cancelar(self):
         self.txt_nome_usuario.setText("")
         self.txt_usuario.setText("")
         self.txt_senha_usuario.setText("")
         self.txt_confirmar_senha.setText("")


    def Adicionar_Trabalho(self):
              
        if not self.txt_nome_trabalho.text():
            self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_trabalho_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_modelo_auto.text():
            self.txt_modelo_auto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_modeloauto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_modelo_auto.setStyleSheet(self.StyleLineEditOk)
        if not self.txt_preco_servico.text():
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_preco_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)       
         
        nome_trabalho = self.txt_nome_trabalho.text()
        ID_funcionario = self.cmb_funcionario_responsavel.itemData(self.cmb_funcionario_responsavel.currentIndex())
        CPF_cliente = self.cmb_cliente.itemData(self.cmb_cliente.currentIndex())
        modelo_automovel = self.txt_modelo_auto.text()
        preco = self.txt_preco_servico.text()

        preco = self.txt_preco_servico.text()

        #Conecta ao Banco e pega a função de Inserir Trabalhos, especificada no Database.py
        db = Database()
        db.connect()

        if getattr(self, 'updating_trabalho', False):
            # Marca a variável _updating_trabalho como False após a atualização
            setattr(self, 'updating_trabalho', False)
            # Chama a função de atualização
            self.update_trabalho(ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco)
        else:
            db.insert_trabalhos(ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Trabalhos")
            msg.setText("Trabalho Adicionado com Sucesso")
            msg.exec_()
            self.txt_nome_trabalho.setText("")
            self.txt_modelo_auto.setText("")
            self.txt_preco_servico.setText("")
            self.lbl_erro_nome_trabalho_vazio.hide()
            self.lbl_erro_modeloauto_vazio.hide()
            self.lbl_erro_preco_vazio.hide()
            self.table_trabalho()

    def atualizar_cpf_cliente(self):
        # Este método é chamado sempre que a combobox é ativada
        index = self.cmb_cliente.currentIndex()
        if index != -1:
            cpf_cliente = self.cmb_cliente.itemData(index)
            print(f"Cliente: {self.cmb_cliente.currentText()}, CPF: {cpf_cliente}")
        else:
            print("Nenhum cliente escolhido")

        # Conectar ao banco de dados
        db = Database()
        db.connect()

    def carregar_nome_na_combobox_cliente(self):
        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar funcionários do banco de dados
        clientes = db.get_cliente()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        self.cmb_cliente.clear()


        # Adicionar funcionários à combobox
        for cliente in clientes:
            cpf_cliente, nome_cliente = cliente  
            self.cmb_cliente.addItem(nome_cliente, userData=cpf_cliente)
            

    def atualizar_id_funcionario(self):
        # Este método é chamado sempre que a combobox é ativada
        index = self.cmb_funcionario_responsavel.currentIndex()
        if index != -1:
            id_funcionario = self.cmb_funcionario_responsavel.itemData(index)
            print(f"Funcionário: {self.cmb_funcionario_responsavel.currentText()}, ID: {id_funcionario}")
        else:
            print("No employee selected")

        # Conectar ao banco de dados
        db = Database()
        db.connect()

    def carregar_usuario_na_combobox_funcionario(self):
        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar funcionários do banco de dados
        funcionarios = db.get_funcionario()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        # Limpar a combobox antes de adicionar novos itens
        self.cmb_funcionario_responsavel.clear()

        # Adicionar funcionários à combobox
        for funcionario in funcionarios:
            id_funcionario, nome_funcionario = funcionario  
            self.cmb_funcionario_responsavel.addItem(nome_funcionario, userData=id_funcionario) 

    def setup_connections_trabalhos_filtro(self):
        self.txt_filtro_trabalhos.textChanged.connect(self.filter_table_trabalhos)

    def filter_table_trabalhos(self, text):
        for row in range(self.tw_trabalhos_geral.topLevelItemCount()):
            item = self.tw_trabalhos_geral.topLevelItem(row)
            nome_index_trabalho = 5 #Coluna onde está o nome do funcionário

            if item.text(nome_index_trabalho).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def table_trabalho(self):

        self.tw_trabalhos.clear()

        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                t.modelo_automovel, t.preco
            FROM trabalhos t
            JOIN funcionario f ON t.ID_funcionario = f.ID
            JOIN cliente c ON t.CPF_cliente = c.CPF
            WHERE t.dia_realizado IS NULL
        """
        result = pd.read_sql_query(query, cn)

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos, [str(row[col]) for col in result.columns])
            item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            # Cria e seta um botão de delete
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            # Obter id_trabalho para esta iteração
            id_trabalho = str(row['id_trabalho'])

            # Conectar o botão de delete à função confirmar_delete_trabalho
            delete_button.clicked.connect(functools.partial(self.excluir_trabalho, id_trabalho))

            # Configurar o botão de delete no item da árvore
            self.tw_trabalhos.setItemWidget(item, self.tw_trabalhos.columnCount() - 1, delete_button)

            # Cria o botão de editar na quarta coluna
            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            # Aqui está a lógica ajustada para o botão "Editar"
            edit_button.clicked.connect(lambda id_trabalho=id_trabalho: self.edit_selected_trabalho(id_trabalho))
            edit_button.clicked.connect(self.carregar_usuario_na_combobox_funcionario)
            edit_button.clicked.connect(self.carregar_nome_na_combobox_cliente)

            self.tw_trabalhos.setItemWidget(item, self.tw_trabalhos.columnCount() - 2, edit_button)

        self.tw_trabalhos.setSortingEnabled(False)

        for i in range(self.tw_trabalhos.columnCount()):
            self.tw_trabalhos.resizeColumnToContents(i)

    def table_trabalhos_geral(self):
        self.tw_trabalhos_geral.clear()
        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                t.modelo_automovel, t.preco, t.dia_realizado
            FROM trabalhos t
            JOIN funcionario f ON t.ID_funcionario = f.ID
            JOIN cliente c ON t.CPF_cliente = c.CPF
        """
        result = pd.read_sql_query(query, cn)

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos_geral, [str(row[col]) for col in result.columns])

        self.tw_trabalhos_geral.setSortingEnabled(False)

        for i in range(self.tw_trabalhos.columnCount()):
            self.tw_trabalhos_geral.resizeColumnToContents(i)

    def table_trabalhos_realizados(self):
        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT id_trabalho, nome_trabalho, id_funcionario, CPF_cliente,
                modelo_automovel, preco, dia_realizado
            FROM trabalhos WHERE dia_realizado IS NOT NULL
        """
        result = pd.read_sql_query(query, cn)

        self.tw_trabalhos_realizados.clear()

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos_realizados, [str(row[col]) for col in result.columns])
            item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

        self.tw_trabalhos_realizados.setSortingEnabled(False)

        for i in range(self.tw_trabalhos_realizados.columnCount()):
            self.tw_trabalhos_realizados.resizeColumnToContents(i)

        self.calcular_soma_precos()

    def excluir_trabalho(self, id_trabalho):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Trabalho", "Deseja realmente excluir o trabalho?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o cliente do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM trabalhos WHERE id_trabalho = ?", (id_trabalho,))
            connection.commit()
            connection.close()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Trabalho")
            msg.setText("Trabalho Excluído Com Sucesso")
            msg.exec_()
            # Atualizar a tabela de trabalhos após a exclusão
            self.table_trabalho()

    def edit_selected_trabalho(self, id_trabalho):
        selected_items = self.tw_trabalhos.selectedItems()

        # Verifica se mais de um item está selecionado
        if len(selected_items) > 1:
            QMessageBox.warning(self, "Alerta", "Selecione apenas um item para editar.")
            return

        # Perguntar ao usuário se eles desejam editar o cliente selecionado
        reply = QMessageBox.question(self, "Editar Trabalho", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o id do trabalho selecionado
            id_trabalho = selected_items[0].text(0)

            # Armazenar o id do trabalho em edição em um atributo da classe
            self.id_trabalho_edicao = id_trabalho

            # Obter os detalhes do trabalho no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            query = """
                SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                    t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                    t.modelo_automovel, t.preco
                FROM trabalhos t
                JOIN funcionario f ON t.ID_funcionario = f.ID
                JOIN cliente c ON t.CPF_cliente = c.CPF
                WHERE t.dia_realizado IS NULL AND t.id_trabalho = ?
            """
            cursor = connection.cursor()
            cursor.execute(query, (id_trabalho,))
            trabalho = cursor.fetchone()
            connection.close()

            self.updating_trabalho = True

            # Definir a página atual como a página de trabalhos
            self.Paginas.setCurrentWidget(self.page_AdicionarTrabalho)
            self.lbl_titulo_cadastro_trabalho.setText("Editar Trabalho")
            self.btn_salvar_trabalho.setText("Editar")

            # Definir o texto dos campos com os detalhes do cliente
            self.txt_nome_trabalho.setText(trabalho[5])

            # Configurar a combobox do funcionário responsável
            if self.cmb_funcionario_responsavel.count() > 0:
                # Utilizar o índice correto (2) para obter o nome do funcionário
                index = self.cmb_funcionario_responsavel.findText(trabalho[2])
                if index != -1:
                    self.cmb_funcionario_responsavel.setCurrentIndex(index)

            # Configurar a combobox do cliente
            if self.cmb_cliente.count() > 0:
                index = self.cmb_cliente.findText(trabalho[4])
                if index != -1:
                    self.cmb_cliente.setCurrentIndex(index)

            self.txt_modelo_auto.setText(trabalho[6])

            self.txt_preco_servico.setText(str(trabalho[7])) 
            self.txt_preco_servico.setText(trabalho[7])



    def update_trabalho(self, ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco):
        # Update do trabalho no database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE trabalhos
            SET ID_funcionario=?, CPF_cliente=?, nome_trabalho=?, modelo_automovel=?, preco=?
            WHERE id_trabalho=?
        """, (ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco, self.id_trabalho_edicao))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Cadastro de Trabalhos")
        msg.setText("Trabalho Editado com Sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_trabalho.clear()
        self.txt_modelo_auto.clear()
        self.txt_preco_servico.clear()
        self.table_trabalho()
        

    def realizar_trabalho(self):

        self.checked_items_out = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items_out.append(child.text(0))

        recurse(self.tw_trabalhos.invisibleRootItem())

        # Pergunta se o usuário realmente deseja fazer isso.
        self.question_trabalhos('realizar')

    def voltar_trabalho(self):

        self.checked_items = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items.append(child.text(0))

        recurse(self.tw_trabalhos_realizados.invisibleRootItem())
        self.question_trabalhos('voltar')   

    def calcular_soma_precos(self):
        # Inicialize a variável para armazenar a soma
        soma_precos_trabalho = 0.0

        # Iterar sobre os itens na árvore
        for item_index in range(self.tw_trabalhos_realizados.topLevelItemCount()):
            item = self.tw_trabalhos_realizados.topLevelItem(item_index)
            preco_str = item.text(5)

            try:
                # Converter o valor do preço para float e somar à variável de soma
                preco = float(preco_str)
                soma_precos_trabalho += preco
            except ValueError:
                # Tratar o caso em que a conversão para float falha (pode ser necessário lidar com isso de acordo com seus requisitos)
                print(f"Erro ao converter preço: {preco_str}")

        self.txt_total_trabalhos.setText(str(soma_precos_trabalho))

    def question_trabalhos(self, table):

        msgBox = QMessageBox()

        if table == 'voltar':
            msgBox.setWindowTitle("Retornar Trabalho")
            msgBox.setText("Deseja retornar os trabalhos selecionados?")
            msgBox.setInformativeText("Os itens selecionados sairão para dos trabalhos feitos \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Trabalhos: {self.checked_items}")

        else:
            msgBox.setWindowTitle("Realizar Trabalho")
            msgBox.setText("Deseja realizar os trabalhos?")
            msgBox.setInformativeText("Os trabalhos sairão dos trabalhos disponíveis\n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Trabalhos: {self.checked_items_out}")

        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()

        if ret == QMessageBox.Yes:
            if table == "voltar":
                self.db = Database()
                print(f"Trabalhos selecionados para voltar: {self.checked_items}")
                self.db.connect()
                self.db.update_trabalho_disponivel(self.checked_items)
                self.db.close_connection()
                self.reset_table()
            else:
                dia_realizado = date.today()
                dia_realizado = dia_realizado.strftime('%d/%m/%Y')
                self.db = Database()
                self.db.connect()
                self.db.UpdateDate_trabalho(dia_realizado, self.checked_items_out)
                self.db.close_connection()
                self.reset_table()


    def CancelarTrabalho(self):
         self.txt_nome_trabalho.setText("")
         self.txt_modelo_auto.setText("")
         self.txt_ID.setText("")
         self.txt_preco_servico.setText("")


    def Adicionar_Cliente(self):
        if not self.txt_nome_cliente.text():
            self.txt_nome_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_cliente_vazio.show()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro do cliente")
            msg.exec_()
            return None
        else:
            self.txt_nome_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_cpf_cliente.text():
            self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cpf_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_telefone_cliente.text():
            self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_telefone_vazio.show()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.date_cliente.text():
            self.date_cliente.setStyleSheet(self.StyleLineEditErro)

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.date_cliente.setStyleSheet(self.StyleLineEditOk)

        # Se não houver erros, você consegue cadastrar ou editar o cliente
        CPF = self.txt_cpf_cliente.text()
        Nome = self.txt_nome_cliente.text()
        Telefone = self.txt_telefone_cliente.text()
        Dia_Agendamento = self.date_cliente.text()
        Peca_Requisitada = self.cmb_pecas.currentText()

        db = Database()
        db.connect()

        if not getattr(self, 'updating_cliente', False):
            # Verifica se o cliente já existe no banco de dados apenas se não estiver atualizando
            cliente_existe = db.check_if_cliente_exists(CPF)

            if cliente_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Cliente Existente")
                msg.setText(f"O cliente com CPF '{CPF}' já existe.")
                msg.exec_()
                self.txt_cpf_cliente.setText("")
                self.txt_nome_cliente.setText("")
                self.txt_telefone_cliente.setText("")
                self.lbl_erro_nome_cliente_vazio.hide()
                self.lbl_erro_cpf_vazio.hide()
                self.lbl_erro_telefone_vazio.hide()

        Chave_Nota = db.get_produto_id_by_descricao(Peca_Requisitada)

        if getattr(self, 'updating_cliente', False):
            # Marca a variável _updating_cliente como False após a atualização
            setattr(self, 'updating_cliente', False)
            # Chama a função de atualização
            self.update_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Nota, Dia_Agendamento)
        else:
            db.insert_cliente(CPF, Nome, Telefone, Dia_Agendamento, Peca_Requisitada, Chave_Nota)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Clientes")
            msg.setText("Cliente Adicionado com Sucesso")
            msg.exec_()
            self.txt_cpf_cliente.setText("")
            self.txt_nome_cliente.setText("")
            self.txt_telefone_cliente.setText("")
            self.lbl_erro_nome_cliente_vazio.hide()
            self.lbl_erro_cpf_vazio.hide()
            self.lbl_erro_telefone_vazio.hide()
            self.table_cliente()


    def atualizar_chave_nota(self):
        # Este método é chamado sempre que a combobox é ativada
        # Atualiza a variável Chave_Nota com base na descrição selecionada
        descricao_selecionada = self.cmb_pecas.currentText()

        db = Database()
        db.connect()
        self.Chave_Nota = db.get_produto_id_by_descricao(descricao_selecionada)
        db.close_connection()
        print(f"Descrição: {descricao_selecionada}, Chave_Nota: {self.Chave_Nota}")

    def carregar_produtos_na_combobox(self):

        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar produtos do banco de dados
        produtos = db.get_produtos()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        # Limpar a combobox antes de adicionar novos itens
        self.cmb_pecas.clear()

       
        for produto in produtos:
            descricao = produto[0] 
            self.cmb_pecas.addItem(descricao)
 
    def table_cliente(self):

        self.tw_cliente.clear()
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()

        # Pega as informações da tabela cliente
        cursor.execute("SELECT CPF, Nome, Telefone, Dia_Agendamento, Peca_Requisitada, Chave_Notas FROM cliente ORDER BY CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento")
        records = cursor.fetchall()

        
        for record in records:
            item = QTreeWidgetItem(self.tw_cliente)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_cliente(item, record))
            self.tw_cliente.setItemWidget(item, 6, delete_button)  # Ajuste na posição

            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.carregar_produtos_na_combobox())
            edit_button.clicked.connect(lambda: self.edit_selected_cliente())
            self.tw_cliente.setItemWidget(item, 7, edit_button)  # Ajuste na posição

        # Resize columns to their contents
        self.tw_cliente.resizeColumnToContents(0)  
        for i in range(1, self.tw_cliente.columnCount()): 
            self.tw_cliente.resizeColumnToContents(i)
            self.tw_cliente.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_cliente.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        # Close the database connection
        connection.close()


    def setup_connections_cliente_filtro(self):
        self.txt_filtro_cliente.textChanged.connect(self.filter_table)

    def filter_table(self, text):
        for row in range(self.tw_cliente.topLevelItemCount()):
            item = self.tw_cliente.topLevelItem(row)
            nome_column_index = 1  

            if item.text(nome_column_index).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def edit_selected_cliente(self):
        selected_items = self.tw_cliente.selectedItems()  # Obter itens selecionados na tabela
        
        # Verifica se mais de um item está selecionado
        if len(selected_items) > 1:
            # Exibe uma mensagem de alerta com ícone de alerta
            QMessageBox.warning(self, "Alerta", "Selecione apenas um item para editar.")
            return
        
        # Perguntar ao usuário se eles desejam editar o cliente selecionado
        reply = QMessageBox.question(self, "Editar Cliente", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o CPF do cliente selecionado
            cpf = selected_items[0].text(0) 

            # Armazenar o CPF do cliente em edição em um atributo da classe
            self.cpf_cliente_edicao = cpf

            # Obter os detalhes do cliente no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM cliente WHERE CPF = ?", (cpf,))
            cliente = cursor.fetchone()
            connection.close()

            # Definir a página atual como a página de clientes
            self.Paginas.setCurrentWidget(self.page_clientes)

            # Definir o texto dos campos com os detalhes do cliente
            self.txt_nome_cliente.setText(cliente[1])
            self.txt_cpf_cliente.setText(cliente[0])
            self.txt_telefone_cliente.setText(cliente[2])

            if self.cmb_pecas.count() > 0:
                index = self.cmb_pecas.findText(cliente[4])
                if index != -1:
                    self.cmb_pecas.setCurrentIndex(index)

                date_string = cliente[3] 
                date = QDate.fromString(date_string, Qt.ISODate)
                self.date_cliente.setDate(date)
                self.updating_cliente = True
                self.lbl_titulo_cadastro_cliente.setText("Atualizar Cliente")
                self.btn_salvar_cliente.setText("Atualizar")


    def update_cliente(self, CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento):
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE cliente
            SET CPF=?, Nome=?, Telefone=?, Peca_Requisitada=?, Chave_Notas=?, Dia_Agendamento=?
            WHERE CPF=?
        """, (CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento, CPF))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Cadastro de Cliente")
        msg.setText("Cliente Editado com Sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_cliente.clear()
        self.txt_telefone_cliente.clear()
        self.date_cliente.clear()
        self.table_cliente()

    def on_btn_salvar_cliente_clicked(self):
        CPF = self.txt_cpf_cliente.text()
        Nome = self.txt_nome_cliente.text()
        Telefone = self.txt_telefone_cliente.text()
        Peca_Requisitada = self.cmb_pecas.currentText()
        Dia_Agendamento = self.date_cliente.text()


        selected_items = self.tw_cliente.selectedItems()
        print("Selected Items:", selected_items)

        if selected_items:
            # Se o cliente estiver selecionado, edita
            CPF = selected_items[0].text(0)

            # Obter a Chave_Notas (anteriormente CPF_Cliente) da peça solicitada
            db = Database()
            db.connect()
            Chave_Notas = db.get_produto_id_by_descricao(Peca_Requisitada)
            db.close_connection()

            self.update_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento)
        else:
            # Se nenhum cliente estiver selecionado, salva normalmente
            self.insert_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento)

    def excluir_cliente(self, CPF):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Cliente", "Deseja realmente excluir o cliente?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o cliente do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM cliente WHERE CPF=?", (CPF,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de fornecedores após a exclusão
            self.table_cliente()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Cliente")
            msg.setText("Cliente Excluído Com Sucesso")
            msg.exec_()

    def confirmar_delete_cliente(self, item, record):
        # Obter o CPF do cliente a ser excluído
        CPF = record[0]

        # Chamar a função para excluir o cliente
        self.excluir_cliente(CPF)


    def eventFilter(self, obj, event):
        if obj is self.txt_cep_fornecedor and event.type() == QEvent.KeyPress and event.key() == Qt.Key_Return:
            self.CepFornecedor()
            return True
        return super().eventFilter(obj, event)

    def CepFornecedor(self):
        cep = self.txt_cep_fornecedor.text()
        resultado = brazilcep.get_address_from_cep(cep)

        if resultado:
            self.txt_logradouro_fornecedor.setText(resultado["street"])
            self.txt_bairro_fornecedor.setText(resultado["district"])
        
            cidade = resultado["city"]
            estado = resultado["uf"]
            self.txt_UF_fornecedor.setText(f"{cidade} / {estado}")
        else:
            print("CEP inválido")


    def Adicionar_fornecedor(self):

        if not self.txt_nome_fantasia.text():
            self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_fantasia_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_razao_social.text():
            self.txt_razao_social.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_razao_social_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_cnpj_fornecedor.text():
            self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cnpj_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_IE_fornecedor.text():
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_IE_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditOk) 

        # Verificar se o preço é um valor numérico
        try:
            IE = float(self.txt_IE_fornecedor.text())
        except ValueError:
            # Se a conversão falhar, exibir mensagem de erro e retornar
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.txt_IE_fornecedor.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("A Inscrição Estadual deve ser um valor numérico.")
            msg.exec_()
            return None
        else:
            # Se a conversão for bem-sucedida, continuar com o cadastro
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)
            self.lbl_erro_preco_vazio.hide()       
        if not self.txt_telefone_fornecedor.text():
            self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_telefone_fornecedor_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_email_fornecedor.text():
            self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_email_fornecedor_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_numero_fornecedor.text():
            self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_numero_casa_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_logradouro_fornecedor.text():
            self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_logradouro_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_bairro_fornecedor.text():
            self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_bairro_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)           

        if not self.txt_UF_fornecedor.text():
            self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cidade_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditOk) 
       

        # Separa cidade e estado
        uf_fornecedor = self.txt_UF_fornecedor.text()
        if '/' in uf_fornecedor:
            cidade, estado = map(str.strip, uf_fornecedor.split('/'))
        else:
            cidade, estado = uf_fornecedor, ''  # Defina estado como vazio

        nome_fantasia = self.txt_nome_fantasia.text()
        razao_social = self.txt_razao_social.text()
        cnpj = self.txt_cnpj_fornecedor.text()
        insc_estadual = self.txt_IE_fornecedor.text()
        telefone = self.txt_telefone_fornecedor.text()
        email = self.txt_email_fornecedor.text()
        cep = self.txt_cep_fornecedor.text()
        numero = self.txt_numero_fornecedor.text()
        logradouro = self.txt_logradouro_fornecedor.text()
        bairro = self.txt_bairro_fornecedor.text()

        db = Database()
        db.connect()
    
        if not getattr(self, 'updating_fornecedor', False):
            # Verifica se o cliente já existe no banco de dados apenas se não estiver atualizando
            fornecedor_existe = db.check_if_fornecedor_exists(cnpj)

            if fornecedor_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Fornecedor Existente")
                msg.setText(f"O fornecedor '{nome_fantasia}' já existe.")
                msg.exec_()
                nome_fantasia = self.txt_nome_fantasia.setText("")
                razao_social = self.txt_razao_social.setText("")
                cnpj = self.txt_cnpj_fornecedor.setText("")
                insc_estadual = self.txt_IE_fornecedor.setText("")
                telefone = self.txt_telefone_fornecedor.setText("")
                email = self.txt_email_fornecedor.setText("")
                cep = self.txt_cep_fornecedor.setText("")
                numero = self.txt_numero_fornecedor.setText("")
                logradouro = self.txt_logradouro_fornecedor.setText("")
                bairro = self.txt_bairro_fornecedor.setText("")

        # Marca a variável _updating_funcionario como False após a atualização
        if getattr(self, 'updating_fornecedor', False):
            setattr(self, 'updating_fornecedor', False)
            cnpj = self.tw_fornecedor.selectedItems()[0].text(0)
            self.update_fornecedor(cnpj, nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado)

        else:
                db.insert_fornecedor(nome_fantasia, razao_social,cnpj, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Cadastro de Fornecedor")
                msg.setText("Fornecedor Adicionado com Sucesso")
                msg.exec_()
                nome_fantasia = self.txt_nome_fantasia.setText("")
                razao_social = self.txt_razao_social.setText("")
                cnpj = self.txt_cnpj_fornecedor.setText("")
                insc_estadual = self.txt_IE_fornecedor.setText("")
                telefone = self.txt_telefone_fornecedor.setText("")
                email = self.txt_email_fornecedor.setText("")
                cep = self.txt_cep_fornecedor.setText("")
                numero = self.txt_numero_fornecedor.setText("")
                logradouro = self.txt_logradouro_fornecedor.setText("")
                bairro = self.txt_bairro_fornecedor.setText("")
                self.table_fornecedor()



    def setup_connections_fornecedor_filtro(self):
        self.txt_filtro_fornecedor.textChanged.connect(self.filter_table_fornecedor)

    def filter_table_fornecedor(self, text):
        for row in range(self.tw_fornecedor.topLevelItemCount()):
            item = self.tw_fornecedor.topLevelItem(row)
            nome_index_fornecedor = 1 #Coluna onde está o nome do funcionário

            if item.text(nome_index_fornecedor).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def table_fornecedor(self):
        self.tw_fornecedor.clear()

        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()

        #Faz uma busca na tabela
        cursor.execute("SELECT cnpj, nome_fantasia, email, telefone FROM fornecedor ORDER BY cnpj, nome_fantasia, email, telefone")
        records = cursor.fetchall()

        # Interage com os records e grava eles na TreeWidget
        for record in records:
            item = QTreeWidgetItem(self.tw_fornecedor)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            # Cria e seta um botão de delete
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                                "background-color: rgb(41, 49, 83); " \
                                "border: 2px solid rgb(55, 100, 183); " \
                                "border-radius: 4px; " \
                                "color: white; " \
                                "}" \
                                "QPushButton:pressed { " \
                                "border: 2px solid rgb(47, 82, 162); " \
                                "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_fornecedor(item, record))
            self.tw_fornecedor.setItemWidget(item, self.tw_fornecedor.columnCount() - 1, delete_button)

            # Cria o botão de editar na quarta coluna
            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                                        "background-color: rgb(41, 49, 83); " \
                                        "border: 2px solid rgb(55, 100, 183); " \
                                        "border-radius: 4px; " \
                                        "color: white; " \
                                        "}" \
                                        "QPushButton:pressed { " \
                                        "border: 2px solid rgb(47, 82, 162); " \
                                        "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.edit_selected_fornecedor())
            self.tw_fornecedor.setItemWidget(item, 4, edit_button)  # Ajuste para a nova posição da coluna

        # Resize columns to their contents
        self.tw_fornecedor.resizeColumnToContents(0) 
        for i in range(1, self.tw_cliente.columnCount()): 
            self.tw_cliente.resizeColumnToContents(i)
            self.tw_cliente.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_cliente.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        # Close the database connection
        connection.close()

    def excluir_fornecedor(self, cnpj):
        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Fornecedor", "Deseja realmente excluir o fornecedor?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o fornecedor do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM fornecedor WHERE cnpj=?", (cnpj,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de fornecedores após a exclusão
            self.table_fornecedor()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Fornecedor")
            msg.setText("Fornecedor excluído com sucesso")
            msg.exec_()

    def confirmar_delete_fornecedor(self, record):
        # Obter o ID do fornecedor a ser excluído
        cnpj = record[0]

        # Chamar a função para excluir o fornecedor
        self.excluir_fornecedor(cnpj)

    def edit_selected_fornecedor(self):
        selected_items = self.tw_fornecedor.selectedItems()  #Pega os itens selecionados na tabela
        if not selected_items:
            return  # Nenhum item selecionado

        # pergunta ao usuário se ele quer editar o fornecedor
        reply = QMessageBox.question(self, "Editar Fornecedor", "Deseja editar o registro selecionado?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Pega o id do fornecedor selecionado
            cnpj = selected_items[0].text(0)  
            # Fetch os detalhes do fornecedor
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM fornecedor WHERE cnpj = ?", (cnpj,))
            fornecedor = cursor.fetchone()
            connection.close()

            if fornecedor:
                #Manda o usuário para a tela de adicionar fornecedor
                self.Paginas.setCurrentWidget(self.page_adicionar_fornecedor)

                #Preenche as textboxes com os dados do fornecedor
                self.txt_nome_fantasia.setText(fornecedor[0]) 
                self.txt_razao_social.setText(fornecedor[1]) 
                self.txt_cnpj_fornecedor.setText(fornecedor[2])
                self.txt_IE_fornecedor.setText(str(fornecedor[3]))
                self.txt_telefone_fornecedor.setText(fornecedor[4])
                self.txt_email_fornecedor.setText(fornecedor[5])
                self.txt_cep_fornecedor.setText(fornecedor[6]) 
                self.txt_numero_fornecedor.setText(fornecedor[7])
                self.txt_logradouro_fornecedor.setText(fornecedor[8]) 
                self.txt_bairro_fornecedor.setText(fornecedor[9])
                self.txt_UF_fornecedor.setText(fornecedor[10] + " / " + fornecedor[11]) 

            self.updating_fornecedor = True
            self.lbl_titulo_fornecedor.setText("Editar Fornecedor")
            self.btn_salvar_fornecedor.setText("Editar")


    def update_fornecedor(self, cnpj, nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado):
        # Faz update do fornecedor no database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE fornecedor
            SET nome_fantasia=?, razao_social=?, insc_estadual=?, telefone=?, email=?, cep=?, numero=?, logradouro=?, bairro=?, cidade=?, estado=?
            WHERE cnpj=?
        """, (nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado, cnpj))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Edição de Fornecedor")
        msg.setText("Fornecedor editado com sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_fantasia.clear()
        self.txt_razao_social.clear()
        self.txt_cnpj_fornecedor.clear()
        self.txt_IE_fornecedor.clear()
        self.txt_telefone_fornecedor.clear()
        self.txt_email_fornecedor.clear()
        self.txt_cep_fornecedor.clear()
        self.txt_numero_fornecedor.clear()
        self.table_fornecedor()


    def Adicionar_Item(self):
        if not self.txt_codigo_produto.text():
            self.txt_codigo_produto.setStyleSheet(self.StyleLineEditErro) 
            self.lbl_erro_codigo_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_codigo_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_quantidade_produto.text():
            self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_qtd_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_descricao_produto.text():
            self.txt_descricao_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_descricao_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_descricao_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_unidade_medida.text():
            self.txt_unidade_medida.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_unidade_medida_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_unidade_medida.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_preco_produto.text():
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_valor_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk) 

        try:
            valorProd = float(self.txt_preco_produto.text())
        except ValueError:
            # Se a conversão falhar, exibir mensagem de erro e retornar
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_preco_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("O preço do produto deve ser um valor numérico.")
            msg.exec_()
            return None   
        else:
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk) 
            self.lbl_erro_preco_produto_vazio.hide()

        NFeRandom = random.randint(1, 100)
        NFe = str(NFeRandom) 
        serieRandom = random.randint(1, 100)
        serie = str(serieRandom)
        data_emissao = "//"
        chaveRandom = random.randint(1,100)
        chave = str(chaveRandom)
        cnpj_emitente = "850000"
        nome_emitente =  "Sigma Systems"
        valorNFe = "58,00"
        itemNotaRandom = random.randint(1, 100)
        itemNota = str(itemNotaRandom)
        cod = self.txt_codigo_produto.text()
        qntd = self.txt_quantidade_produto.text()
        descricao = self.txt_descricao_produto.text()
        unidade_medida = self.txt_unidade_medida.text()
        valorProd = self.txt_preco_produto.text()
        data_importacao = date.today()
        data_importacao = data_importacao.strftime('%d/%m/%Y')
        data_saida = ""

        db = Database()
        db.connect()

        produto_existe = db.check_if_produto_exists(cod)

        if produto_existe:
            db.close_connection()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Produto Existente")
            msg.setText(f"O Produto com código'{cod}' já existe.")
            msg.exec_()
            self.txt_codigo_produto.setText("")
            self.txt_quantidade_produto.setText("")
            self.txt_descricao_produto.setText("")
            self.txt_unidade_medida.setText("")
            self.txt_preco_produto.setText("")
            self.lbl_erro_codigo_produto_vazio.hide()
            self.lbl_erro_qtd_produto_vazio.hide()
            self.lbl_erro_descricao_produto_vazio.hide()
            self.lbl_erro_unidade_medida_vazio.hide()
            self.lbl_erro_preco_produto_vazio.hide()
        else:
            db.inserir_produtos(NFe, serie, data_emissao, chave, cnpj_emitente, nome_emitente,
                                valorNFe, itemNota, cod, qntd, descricao, unidade_medida, valorProd,
                                data_importacao, data_saida)
            db.close_connection()
            self.reset_table()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Produtos")
            msg.setText("Produto adicionado com sucesso")
            msg.exec()

            self.txt_codigo_produto.setText("")
            self.txt_quantidade_produto.setText("")
            self.txt_descricao_produto.setText("")
            self.txt_unidade_medida.setText("")
            self.txt_preco_produto.setText("")
            self.lbl_erro_codigo_produto_vazio.hide()
            self.lbl_erro_qtd_produto_vazio.hide()
            self.lbl_erro_descricao_produto_vazio.hide()
            self.lbl_erro_unidade_medida_vazio.hide()
            self.lbl_erro_preco_produto_vazio.hide()


    def CancelarProduto(self):
         self.txt_codigo_produto.setText("")
         self.txt_quantidade_produto.setText("")
         self.txt_descricao_produto.setText("")
         self.txt_unidade_medida.setText("")
         self.txt_preco_produto.setText("")


    def open_path(self):
        self.path = QFileDialog.getExistingDirectory(self,str("Open Directory"),
                                                            "/home",
                                                            QFileDialog.ShowDirsOnly
                                                            | QFileDialog.DontResolveSymlinks)
        self.txt_arquivo_xml.setText(self.path)
    
    def import_arquivos_xml(self):

        xml = Read_xml(self.path)
        all = xml.all_files()

        db = Database()
        db.connect()

        if len(all) == 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Importação XML")
            msg.setText("Não foi possível concluir a importação. A pasta selecionada não possui um arquivo XML NF-e.")
            msg.exec_()
            return

        # try/except para capturar erros
        try:
            for i in all:
                fullDataSet = xml.nfe_data(i)
                db.insert_data(fullDataSet)
        except sqlite3.IntegrityError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Importação XML")
            msg.setText("Nota já existe no banco")
            msg.exec_()
        
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Importação XML")
        msg.setText("Importação Concluída")
        msg.exec_()
        self.reset_table()
        db.close_connection()
            

    def table_estoque(self):
        cn  =  sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("SELECT * FROM Notas WHERE data_saida = ''", cn)
        result = result.values.tolist()

        self.x = ""

        for i in result:
            #faz o check para identificar a mesma nota e adicionar um nivel
            if i[0] == self.x:
                QTreeWidgetItem(self.campo, i)
            else:
                self.campo = QTreeWidgetItem(self.tw_estoque, i)
                self.campo.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            self.x = i[0]

        self.tw_estoque.setSortingEnabled(False)
        self.tw_estoque.itemChanged.connect(self.botões)
   
        for i in range(1,14):
            self.tw_estoque.resizeColumnToContents(i)


    def botões(self):
        # Inicialize a lista de itens selecionados
        self.checked_items_out = []
    
        # Chame a função recursiva para verificar os estados de verificação
        self.recurse(self.tw_estoque.invisibleRootItem())
    
        # Verifique se a lista de itens selecionados não está vazia para mostrar os botões
        if self.checked_items_out:
            self.btn_excluir.show()
        else:
            self.btn_excluir.hide()

    def recurse(self, parent_item):
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            grand_children = child.childCount()
            if grand_children > 0:
                self.recurse(child)
            if child.checkState(0) == QtCore.Qt.Checked:
                self.checked_items_out.append(child.text(0))

    def questionExcluir(self):
        msgBox = QMessageBox()
        msgBox.setText("Deseja excluir os itens selecionados?")
        msgBox.setInformativeText("Os itens selecionados sairão do estoque para sempre.\nClique em 'Yes' para confirmar.")
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDetailedText(f"Notas: {', '.join(self.checked_items_out)}")
        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()
    
        if ret == QMessageBox.Yes:
            self.db = Database()
            self.db.connect()
            self.db.excluir_itens(self.checked_items_out)
            self.db.close_connection()
            self.reset_table()
        else:
            self.db.close_connection()


    def table_saida(self):
        cn  =  sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("""SELECT NFe, serie, data_importacao, data_saida
         FROM Notas WHERE data_saida != ''""", cn)
        result = result.values.tolist()

        self.x = ""

        for i in result:
            #faz o check para identificar a mesma nota e adicionar um nivel
            if i[0] == self.x:
                QTreeWidgetItem(self.campo, i)
            else:
                self.campo = QTreeWidgetItem(self.tw_saida, i)
                self.campo.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            self.x = i[0]

        self.tw_saida.setSortingEnabled(False)
   
        for i in range(1,14):
            self.tw_saida.resizeColumnToContents(i)


    def table_geral(self):
        # Connect to the SQLite database
        db = QSqlDatabase("QSQLITE")
        db.setDatabaseName("AutoCenter.db")
        db.open()

        # Create a QSqlTableModel
        self.model = QSqlTableModel(db=db)

        # Set the table name
        self.model.setTable("Notas")

        # Select all data from the table
        self.model.select()

        # Set the model to the QTableView
        self.tb_geral.setModel(self.model)

        # Set the table headers
        self.tb_geral.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)


    def reset_table(self):
        self.tw_estoque.clear()
        self.tw_saida.clear()
        self.tw_cliente.clear()
        self.tw_fornecedor.clear()
        self.tw_trabalhos_realizados.clear()
        self.tw_trabalhos.clear()
        self.tw_trabalhos_geral.clear()

        self.table_saida()
        self.table_estoque()
        self.table_cliente()
        self.table_geral()
        self.table_trabalhos_geral()
        self.table_funcionario()
        self.table_fornecedor()
        self.table_trabalho()
        self.table_trabalhos_realizados()


    def update_filter(self, s):
        s = re.sub("[\W_]+", "", s)
        filter_str = 'descricao LIKE "%{}%"'.format(s)
        self.model.setFilter(filter_str)


    def gerar_saida(self):

        self.checked_items_out = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items_out.append(child.text(0))
    
        recurse(self.tw_estoque.invisibleRootItem())

        #Pergunta se usuario realmente deseja fazer isso.
        self.question('saída')


    def gerar_estorno(self):
        
        self.checked_items = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items.append(child.text(0))

        recurse(self.tw_saida.invisibleRootItem())
        self.question('estorno')      


    def question(self, table):

        msgBox = QMessageBox()

        if table == 'estorno':
            msgBox.setText("Deseja estornar as notas selecionadas?")
            msgBox.setInformativeText("Os itens selecionados voltarão para o estoque \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Notas: {self.checked_items}")

        else:
            msgBox.setText("Deseja Gerar saída das nota selecionadas?")
            msgBox.setInformativeText("As notas abaixo sairão do estoque \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Notas: {self.checked_items_out}")

        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()

        if ret == QMessageBox.Yes:
            if table == "estorno":
                self.db = Database()
                self.db.connect()
                self.db.update_estorno(self.checked_items)
                self.db.close_connection()
                self.reset_table()
            else:
                data_saida = date.today()
                data_saida = data_saida.strftime('%d/%m/%Y')
                self.db = Database()
                self.db.connect()
                self.db.UpdateDate_estoque(data_saida, self.checked_items_out)
                self.db.close_connection()
                self.reset_table()

    
    def Converter_Excel(self):
        cnx = sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("SELECT * FROM Notas", cnx)
        
        # Definir o caminho para a pasta raiz do projeto
        project_root = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(project_root, "Notas.xlsx")
        
        # Salvar o arquivo Excel
        result.to_excel(file_path, sheet_name='Notas', index=False)

        # Parte 2: Aplicar estilização ao arquivo Excel gerado
        self.estilizar_tabela_excel(file_path)

        # Mensagem de sucesso com ícone
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Notas em XML")
        msg.setText("Relatório gerado com sucesso!")
        
        # Adicionar ícone à QMessageBox
        icon_path = os.path.join(project_root, "Imagens", "excel.png")
        msg.setWindowIcon(QIcon(icon_path))


        msg.exec_()

    def estilizar_tabela_excel(self, file_path):
        # Abrir o arquivo Excel
        print("Iniciando estilização da tabela...")
        wb = load_workbook(file_path)
        ws = wb.active

        # Configurar um estilo personalizado para o cabeçalho
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Arial', color='FFFFFFFF', bold=True)  # Fonte Arial, cor branca para o texto
        header_style.fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')  # Cor ARGB para o preenchimento
        header_style.alignment = Alignment(horizontal='center', vertical='center')  # Centralizar texto

        thin_border = Border(left=Side(style='thin', color='FF6885bf'),
                            top=Side(style='thin', color='FF6885bf'),
                            right=Side(style='thin', color='FF6885bf'),
                            bottom=Side(style='thin', color='FF6885bf'))

        header_style.border = thin_border

        # Configurar um estilo personalizado para os campos
        cell_style = NamedStyle(name="cell_style")
        cell_style.font = Font(name='Arial', color='FFFFFFFF', bold=True)  # Fonte Arial, cor branca para o texto
        cell_style.fill = PatternFill(start_color='FF071325', end_color='FF071325', fill_type='solid')  # Cor ARGB para o preenchimento
        cell_style.border = thin_border
        cell_style.alignment = Alignment(horizontal='center', vertical='center')  # Centralizar texto

        # Aplicar o estilo ao cabeçalho da tabela
        for col_num, col in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1), start=1):
            for header_cell in col:
                header_cell.style = header_style

        # Alterar a formatação das células (exceto o cabeçalho)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            for cell in row:
                cell.style = cell_style

        # Autoajustar largura das colunas e altura das linhas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                try:
                    if len(str(cell.value)) > max_height:
                        max_height = len(cell.value)
                except:
                    pass
            adjusted_height = (max_height + 2)
            ws.row_dimensions[row[0].row].height = adjusted_height

        # Salvar as alterações
        wb.save(file_path)

    def onPrint(self):
        self.printDialog()


    def printDialog(self):

        filePath, filter = QFileDialog.getOpenFileName(self, 'Open file', '', 'PDF (*.pdf)')
        if not filePath:
            return
        file_extension = os.path.splitext(filePath)[1]

        if file_extension == ".pdf":
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                with tempfile.TemporaryDirectory() as path:
                    images = convert_from_path(filePath, dpi=300, output_folder=path, poppler_path=r'C:\Program Files\poppler-0.68.0\bin')
                    painter = QPainter()
                    painter.begin(printer)
                    for i, image in enumerate(images):
                        if i > 0:
                            printer.newPage()
                        rect = painter.viewport()
                        qtImage = ImageQt(image)
                        qtImageScaled = qtImage.scaled(rect.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        painter.drawImage(rect, qtImageScaled)
                    painter.end()
        else:
            pass

    def graphic(self):

        cnx = sqlite3.connect("AutoCenter.db")
        estoque = pd.read_sql_query('SELECT * FROM Notas', cnx)
        saida = pd.read_sql_query("SELECT * FROM Notas WHERE data_saida != ''", cnx)

        estoque = len(estoque)
        saida = len(saida)

        labels = "Estoque", "Saídas"
        sizes = [estoque, saida]
        fig1, axl = plt.subplots()
        axl.pie(sizes, labels=labels, autopct='%1.1f%%', shadow=True, startangle=90)
        axl.axis("equal")

        plt.show()


class MainWindowUsuario(QMainWindow, Ui_MainWindowUsuario):
    def __init__(self):
        super(MainWindowUsuario, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Sistema de Gerenciamento")
        self.showFullScreen()
        appIcon = QIcon('Imagens/Estatua.png')
        self.setWindowIcon(appIcon)
        #MudaAsPáginas do Sistema
        self.btn_menu.clicked.connect(lambda: self.moverMenuLateral())
        self.btn_menu_fornecedores.clicked.connect(lambda: self.Reset_Fornecedores ())
        self.btn_menu_nota_fiscal.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.Page_Importar_XML))
        self.btn_confUser.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_usuarios))
        self.btn_adicionar_cliente.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_clientes))
        self.btn_adicionar_item.clicked.connect(lambda: self.Reset_Produtos())

        self.btn_menu_estoque.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_Estoque))
        self.btn_menu_fornecedores.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_fornecedor))
        self.btn_adicionar_fornecedor.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_adicionar_fornecedor))
        self.btn_adicionar_fornecedor.clicked.connect(lambda: self.Reset_Fornecedores())
        self.btn_sair.clicked.connect(lambda: self.Sair())
        self.btn_menu_trabalho.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_trabalhos))
        self.btn_adicionar_trabalho.clicked.connect(lambda: self.Reset_Trabalhos())
        self.btn_adicionar_trabalho.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.page_AdicionarTrabalho))


        #Botões Cadastro Cliente
        self.btn_cancelar_cadastro.clicked.connect(lambda: self.Cancelar())
        self.btn_salvar_cadastro.clicked.connect(lambda: self.cadastrar_usuario()) 
        self.btn_olho_senha.clicked.connect(lambda:self.toggleVisibility())
        self.btn_olho_senha_confirmar_senha.clicked.connect(lambda:self.toggleVisibility2())

        #Botões Adicionar Trabalho
        self.btn_cancelar_trabalho.clicked.connect(lambda: self.CancelarTrabalho())
        self.btn_salvar_trabalho.clicked.connect(lambda: self.Adicionar_Trabalho())

        #Botões Adicionar Produtos
        self.btn_salvar_produto.clicked.connect(lambda: self.Adicionar_Item())
        self.btn_cancelar_produto.clicked.connect(lambda: self.CancelarProduto())

        #Botões Adicionar Clientes
        self.btn_salvar_cliente.clicked.connect(lambda: self.Adicionar_Cliente())

        # Botao Salvando Conexao Banco de Dados
        self.txt_cep_fornecedor.returnPressed.connect(self.CepFornecedor)
        self.btn_salvar_fornecedor.clicked.connect(self.Adicionar_fornecedor)
        self.btn_adicionar_fornecedor.clicked.connect(self.Reset_Fornecedores())

        # Instala um filtro de eventos no campo de texto
        self.txt_cep_fornecedor.installEventFilter(self)

        #Esconder Labels de Erro
        self.lbl_erro_nome_vazio.hide()
        self.lbl_erro_usuario_vazio.hide()
        self.lbl_erro_senha_vazio.hide()
        self.lbl_erro_confirmar_senha_vazio.hide()
        self.lbl_erro_nome_trabalho_vazio.hide()
        self.lbl_erro_modeloauto_vazio.hide()
        self.lbl_erro_preco_vazio.hide()

        #Esconder Botões
        self.btn_excluir.hide()

        #Arquivo XML
        self.btn_abrir_diretorio.clicked.connect(self.open_path)
        self.btn_importar_xml.clicked.connect(self.import_arquivos_xml)

        #tabelas
        self.btn_gerar_saida.clicked.connect(self.gerar_saida)
        self.btn_gerar_estorno.clicked.connect(self.gerar_estorno)
        self.btn_realizar_trabalho.clicked.connect(self.realizar_trabalho)
        self.btn_voltar_trabalho.clicked.connect(self.voltar_trabalho)
        self.reset_table()

        #Botôes Estoque
        self.btn_adicionar_item.clicked.connect(lambda: self.Paginas.setCurrentWidget(self.Page_Produto))
        self.btn_gerar_excel.clicked.connect(lambda: self.Converter_Excel())
        self.btn_grafico.clicked.connect(lambda: self.graphic())
        self.btn_imprimir.clicked.connect(lambda: self.onPrint())
        self.btn_excluir.clicked.connect(lambda: self.questionExcluir())

        self.btn_adicionar_cliente.clicked.connect(self.carregar_produtos_na_combobox)
        self.btn_adicionar_trabalho.clicked.connect(self.carregar_usuario_na_combobox_funcionario)
        self.btn_adicionar_trabalho.clicked.connect(self.carregar_nome_na_combobox_cliente)


        # ComboBoxes
        self.cmb_pecas.currentTextChanged.connect(self.atualizar_id_funcionario)
        self.cmb_cliente.currentTextChanged.connect(self.atualizar_cpf_cliente)
        self.cmb_funcionario_responsavel.currentTextChanged.connect(self.atualizar_id_funcionario)
        self.Chave_Nota = None

        #Updates
        self.updating_cliente = False
        self.updating_funcionario = False
        self.updating_fornecedor = False
        self.updating_trabalho = False

        #Filtros
        self.setup_connections_cliente_filtro()  #configura a filtragem
        self.setup_connections_funcionario_filtro()
        self.setup_connections_trabalhos_filtro()
        self.setup_connections_fornecedor_filtro()
        self.txt_filtro.textChanged.connect(self.update_filter)

        #Cancelar
        self.btn_cancelar_cliente.clicked.connect(self.Cancelar_Cliente)
        self.btn_cancelar_fornecedor.clicked.connect(self.Cancelar_Fornecedores)
        self.btn_cancelar_trabalho.clicked.connect(self.Cancelar_Trabalho)


    #Estilos
    StyleLineEditOk = ("QLineEdit{\n"
"    border: 2px solid rgb(47, 82, 162);\n"
"    border-radius: 5px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")

    StyleLineEditErro = ("QLineEdit{\n"
"    border: 2px solid rgb(255, 21, 21);\n"
"    border-radius: 5px;\n"
"    background-color: rgb(30,30,30);\n"
"    color: rgb(100, 100, 100, 100);\n"
"}\n"
"QLineEdit{\n"
"    color: rgb(200, 200, 200);\n"
"}")


    def Sair(self):
                   
            self.w = Login()
            self.w.show()
            self.close()


    def ApagarLabelErros(self):
                   
        self.lbl_erro_endereco_vazio.hide()
        self.lbl_erro_nome_bd_vazio.hide()
        self.lbl_erro_usuario_bd_vazio.hide()
        self.lbl_erro_senha_bd_vazio.hide()

   
    def Reset_Funcionario(self):
        self.lbl_titulo_cadastro_usuario.setText("Cadastro de Funcionarios")
        self.btn_salvar_cadastro.setText("Salvar")
        self.txt_nome_usuario.setText("")
        self.txt_usuario.setText("")
        self.txt_senha_usuario.setText("")
        self.txt_confirmar_senha.setText("")
        self.txt_nome_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_senha_usuario.setStyleSheet(self.StyleLineEditOk)
        self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_vazio.hide()
        self.lbl_erro_usuario_vazio.hide()
        self.lbl_erro_senha_vazio.hide()
        self.lbl_erro_confirmar_senha_vazio.hide()

    def Reset_Clientes(self):
        self.lbl_titulo_cadastro_cliente.setText("Cadastro de Clientes")
        self.btn_salvar_cliente.setText("Salvar")
        self.txt_cpf_cliente.setText("")
        self.txt_nome_cliente.setText("")
        self.txt_telefone_cliente.setText("")
        self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditOk)
        self.txt_nome_cliente.setStyleSheet(self.StyleLineEditOk)
        self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_cliente_vazio.hide()
        self.lbl_erro_cpf_vazio.hide()
        self.lbl_erro_telefone_vazio.hide()
        self.lbl_erro_telefone_vazio.hide()

    def Reset_Trabalhos(self):
        self.lbl_titulo_cadastro_trabalho.setText("Cadastro de Trabalhos")
        self.btn_salvar_trabalho.setText("Salvar")
        self.txt_nome_trabalho.setText("")
        self.txt_modelo_auto.setText("")
        self.txt_preco_servico.setText("")
        self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditOk)
        self.txt_modelo_auto.setStyleSheet(self.StyleLineEditOk)
        self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_trabalho_vazio.hide()
        self.lbl_erro_modeloauto_vazio.hide()
        self.lbl_erro_preco_vazio.hide()

    def Reset_Fornecedores(self):
        self.lbl_titulo_fornecedor.setText("Fornecedores")
        self.btn_salvar_fornecedor.setText("Salvar")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_IE_fornecedor.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_email_fornecedor.setText("")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_telefone_fornecedor.setText("")
        self.txt_cep_fornecedor.setText("")
        self.txt_numero_fornecedor.setText("")
        self.txt_logradouro_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")
        self.txt_UF_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")    
        self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)
        self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)
        self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)
        self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)
        self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_cep_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_nome_fantasia_vazio.hide()
        self.lbl_erro_razao_social_vazio.hide()
        self.lbl_erro_cnpj_vazio.hide()
        self.lbl_erro_IE_vazio.hide()
        self.lbl_erro_telefone_fornecedor_vazio.hide()
        self.lbl_erro_email_fornecedor_vazio.hide()
        self.lbl_erro_numero_casa_vazio.hide()
        self.lbl_erro_logradouro_vazio.hide()
        self.lbl_erro_bairro_vazio.hide()
        self.lbl_erro_cidade_vazio.hide()

    def Reset_Produtos(self):
        self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_unidade_medida.setStyleSheet(self.StyleLineEditOk)
        self.txt_descricao_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditOk)
        self.txt_codigo_produto.setStyleSheet(self.StyleLineEditOk)
        self.lbl_erro_qtd_produto_vazio.hide()
        self.lbl_erro_codigo_produto_vazio.hide()
        self.lbl_erro_descricao_produto_vazio.hide()
        self.lbl_erro_unidade_medida_vazio.hide()
        self.lbl_erro_preco_produto_vazio.hide()

    def Cancelar_Cliente(self):
        self.txt_cpf_cliente.setText("")
        self.txt_nome_cliente.setText("")
        self.txt_telefone_cliente.setText("")

    def Cancelar_Fornecedores(self):
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_IE_fornecedor.setText("")
        self.txt_cnpj_fornecedor.setText("")
        self.txt_email_fornecedor.setText("")
        self.txt_nome_fantasia.setText("")
        self.txt_razao_social.setText("")
        self.txt_telefone_fornecedor.setText("")
        self.txt_cep_fornecedor.setText("")
        self.txt_numero_fornecedor.setText("")
        self.txt_logradouro_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")
        self.txt_UF_fornecedor.setText("")
        self.txt_bairro_fornecedor.setText("")    

    def Cancelar_Trabalho(self):
        self.txt_nome_trabalho.setText("")
        self.txt_modelo_auto.setText("")
        self.txt_preco_servico.setText("")


    def keyPressEvent(self, event: QKeyEvent):
        
        if event.key() == Qt.Key_F11: 
            if self.isFullScreen(): 
                self.showNormal() 
            else: self.showFullScreen() 
        elif event.key() == Qt.Key_Escape and self.isFullScreen(): 
            self.showNormal()


    def moverMenuLateral(self):
         width = self.Left_Menu.width()

         # Se estiver minimizado
         if width == 80:
              #Expandir Menu
              newWidth = 250
         else:
              #VoltaAoNormal
              newWidth = 80

        #Animar Transição
         self.animation = QtCore.QPropertyAnimation (self.Left_Menu, b"maximumWidth")
         self.animation.setDuration(250)
         self.animation.setStartValue(width)
         self.animation.setEndValue(newWidth)
         self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
         self.animation.start()


    def toggleVisibility(self):
        if self.txt_senha_usuario.echoMode()==QLineEdit.Normal:
            self.txt_senha_usuario.setEchoMode(QLineEdit.Password)
        else:
            self.txt_senha_usuario.setEchoMode(QLineEdit.Normal)
            
    def toggleVisibility2(self):
        if self.txt_confirmar_senha.echoMode()==QLineEdit.Normal:
            self.txt_confirmar_senha.setEchoMode(QLineEdit.Password)
        else:
            self.txt_confirmar_senha.setEchoMode(QLineEdit.Normal)


    def cadastrar_usuario(self):
         
        #Checka se as senhas são iguais
        if self.txt_senha_usuario.text() != self.txt_confirmar_senha.text():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Senhas Conflitantes")
            msg.setText("Senhas Divergentes")
            msg.exec_()
            return None
         
        if not self.txt_nome_usuario.text():
            self.txt_nome_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_usuario.text():
            self.txt_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_usuario_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_senha_usuario.text():
            self.txt_senha_usuario.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_senha_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_senha_usuario.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_confirmar_senha.text():
            self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_confirmar_senha_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_confirmar_senha.setStyleSheet(self.StyleLineEditOk)              

       
        Nome = self.txt_nome_usuario.text()
        Usuario = self.txt_usuario.text()
        Senha = self.txt_senha_usuario.text()
        Acesso = self.cmb_perfil.currentText()

        db = Database()
        db.connect()


        if not getattr(self, 'updating_funcionario', False):
            # Verifica se o funcionario já existe no banco de dados apenas se não estiver atualizando
            funcionario_existe = db.check_if_user_exists(Usuario)

            if funcionario_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Funcionário Existente")
                msg.setText(f"O funcionário '{Usuario}' já existe.")
                msg.exec_()
                self.txt_cpf_cliente.setText("")
                self.txt_nome_cliente.setText("")
                self.txt_telefone_cliente.setText("")
                self.lbl_erro_nome_cliente_vazio.hide()
                self.lbl_erro_cpf_vazio.hide()
                self.lbl_erro_telefone_vazio.hide()

        # Marca a variável _updating_funcionario como False após a atualização
        if getattr(self, 'updating_funcionario', False):
            setattr(self, 'updating_funcionario', False)
            ID = self.tw_funcionario.selectedItems()[0].text(0)  # Assuming the ID is in the first column
            self.update_funcionario(ID, Nome, Usuario, Senha, Acesso)

        else:
            length_ok = len(Senha) >= 8
            uppercase_ok = any(c.isupper() for c in Senha)
            lowercase_ok = any(c.islower() for c in Senha)
            digit_ok = any(c.isdigit() for c in Senha)
            special_ok = any(not c.isalnum() for c in Senha)
            error_messages = []

            if not length_ok:
                error_messages.append("A senha requer 8 caracteres")
            if not uppercase_ok:
                error_messages.append("A senha requer letras maiúsculas")
            if not lowercase_ok:
                error_messages.append("A senha requer letras minúsculas")
            if not digit_ok:
                error_messages.append("A senha requer pelo menos um número")
            if not special_ok:
                error_messages.append("A senha requer pelo menos um caractere especial")

            if error_messages:
                strength = "Fraca" # Ou outra força adequada, dependendo dos erros
                self.lbl_erro_senha_vazio.show()
                error_message = f"Resultado da verificação: Senha {strength}\n" + "\n".join(error_messages)
                self.lbl_erro_senha_vazio.setText(error_message)
                self.txt_senha_usuario.setStyleSheet(self.StyleLineEditErro)
            else:
                # Se não houver erros, você consegue cadastrar o usuário
                strength = "Muito Forte"
                db.insert_funcionario(Nome, Usuario, Senha, Acesso)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Cadastro de Funcionários")
                msg.setText("Funcionário Adicionado com Sucesso")
                msg.exec_()
                self.txt_nome_usuario.setText("")
                self.txt_usuario.setText("")
                self.txt_senha_usuario.setText("")
                self.txt_confirmar_senha.setText("")
                self.lbl_erro_nome_vazio.hide()
                self.lbl_erro_usuario_vazio.hide()
                self.lbl_erro_senha_vazio.hide()
                self.lbl_erro_confirmar_senha_vazio.hide()
                self.table_funcionario()


    def table_funcionario(self):
        # Clear the existing items in the table
        self.tw_funcionario.clear()

        # Create a database connection
        connection = sqlite3.connect("AutoCenter.db")
        # Create a cursor object
        cursor = connection.cursor()

        # Pega as informações da tabela funcionário
        cursor.execute("SELECT * FROM funcionario")
        records = cursor.fetchall()

        # Iterate over the records and add them to the tree widget
        for record in records:
            item = QTreeWidgetItem(self.tw_funcionario)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            # Create and set the delete button for the sixth column
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_funcionario(item, record))
            self.tw_funcionario.setItemWidget(item, 5, delete_button)  # Ajuste na posição

            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.edit_selected_funcionario())
            edit_button.clicked.connect(lambda: self.ApagarLabelErros())
            self.tw_funcionario.setItemWidget(item, 6, edit_button)  # Ajuste na posição

        # Resize columns to their contents
        self.tw_funcionario.resizeColumnToContents(0)  # Resize the first column for check states
        for i in range(1, self.tw_cliente.columnCount()):  # Resize remaining columns
            self.tw_funcionario.resizeColumnToContents(i)
            self.tw_funcionario.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_funcionario.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        # Close the database connection
        connection.close()


    def setup_connections_funcionario_filtro(self):
        self.txt_filtro_funcionario.textChanged.connect(self.filter_table_funcionario)

    def filter_table_funcionario(self, text):
        for row in range(self.tw_funcionario.topLevelItemCount()):
            item = self.tw_funcionario.topLevelItem(row)
            nome_index_funcionario = 1 #Coluna onde está o nome do funcionário

            if item.text(nome_index_funcionario).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def edit_selected_funcionario(self):
        selected_items = self.tw_funcionario.selectedItems()  # Obter itens selecionados na tabela
        
        # Perguntar ao usuário se eles desejam editar o funcionario selecionado
        reply = QMessageBox.question(self, "Editar Funcionário", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o ID do funcionario selecionado
            ID = selected_items[0].text(0) 

            # Obter os detalhes do Funcionario no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM funcionario WHERE ID = ?", (ID,))
            funcionario = cursor.fetchone()
            connection.close()

            # Definir a página atual como a página de funcionários
            self.Paginas.setCurrentWidget(self.Page_CadastroUsuario)

            # Definir o texto dos campos com os detalhes do funcionário
            self.txt_nome_usuario.setText(funcionario[1])
            self.txt_usuario.setText(funcionario[2])
            self.txt_senha_usuario.setText(funcionario[3])
            self.txt_confirmar_senha.setText(funcionario[3])
            self.btn_salvar_cadastro.setText("Editar")
            self.lbl_titulo_cadastro_usuario.setText("Editar Funcionario")


            if self.cmb_perfil.count() > 0:
                index = self.cmb_perfil.findText(funcionario[4])
                if index != -1:
                    self.cmb_perfil.setCurrentIndex(index)

            self.updating_funcionario = True


    def update_funcionario(self, ID, Nome, Usuario, Senha, Acesso):
        # Update the funcionario in the database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE funcionario
            SET Nome=?, Usuario=?, Senha=?, Acesso=?
            WHERE ID=?
        """, (Nome, Usuario, Senha, Acesso, ID))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Edição de Funcionário")
        msg.setText("Funcionário Editado com Sucesso")
        msg.exec_()

        # Clear input fields
        self.txt_nome_usuario.clear()
        self.txt_usuario.clear()
        self.txt_senha_usuario.clear()
        self.txt_confirmar_senha.clear()
        self.table_funcionario()

    def on_btn_salvar_funcionario_clicked(self):
        # Get the values from the text boxes
        Nome = self.txt_nome_usuario.text()
        Usuario = self.txt_usuario.text()
        Senha = self.txt_senha_usuario.text()
        Acesso = self.cmb_perfil.currentText()

        # Checa se o funcionário está selecionado
        selected_items = self.tw_funcionario.selectedItems()
        print("Selected Items:", selected_items)

        if selected_items:
            # Se o funcionario estiver selecionado, edita
            ID = selected_items[0].text(0)

            db = Database()
            db.connect()
            db.close_connection()

            self.update_funcionario()
        else:
            # Se nenhum funcionario estiver selecionado, salva normalmente
            db.insert_funcionario(Nome, Usuario, Senha, Acesso)

    def excluir_funcionario(self, ID):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Funcionário", "Deseja realmente excluir o funcionário?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o funcionario do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM funcionario WHERE ID=?", (ID,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de funcionários após a exclusão
            self.table_funcionario()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Funcionário")
            msg.setText("Funcionário Excluído Com Sucesso")
            msg.exec_()

    def confirmar_delete_funcionario(self, item, record):

        # Obter o ID do funcionario a ser excluído
        ID = record[0]

        # Chamar a função para excluir o funcionario
        self.excluir_funcionario(ID)


    def Cancelar(self):
         self.txt_nome_usuario.setText("")
         self.txt_usuario.setText("")
         self.txt_senha_usuario.setText("")
         self.txt_confirmar_senha.setText("")


    def Adicionar_Trabalho(self):
              
        if not self.txt_nome_trabalho.text():
            self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_trabalho_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_trabalho.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_modelo_auto.text():
            self.txt_modelo_auto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_modeloauto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_modelo_auto.setStyleSheet(self.StyleLineEditOk)
        if not self.txt_preco_servico.text():
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_preco_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)       
         
        nome_trabalho = self.txt_nome_trabalho.text()
        ID_funcionario = self.cmb_funcionario_responsavel.itemData(self.cmb_funcionario_responsavel.currentIndex())
        CPF_cliente = self.cmb_cliente.itemData(self.cmb_cliente.currentIndex())
        modelo_automovel = self.txt_modelo_auto.text()
        preco = self.txt_preco_servico.text()

        preco = self.txt_preco_servico.text()

        #Conecta ao Banco e pega a função de Inserir Trabalhos, especificada no Database.py
        db = Database()
        db.connect()

        if getattr(self, 'updating_trabalho', False):
            # Marca a variável _updating_trabalho como False após a atualização
            setattr(self, 'updating_trabalho', False)
            # Chama a função de atualização
            self.update_trabalho(ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco)
        else:
            db.insert_trabalhos(ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Trabalhos")
            msg.setText("Trabalho Adicionado com Sucesso")
            msg.exec_()
            self.txt_nome_trabalho.setText("")
            self.txt_modelo_auto.setText("")
            self.txt_preco_servico.setText("")
            self.lbl_erro_nome_trabalho_vazio.hide()
            self.lbl_erro_modeloauto_vazio.hide()
            self.lbl_erro_preco_vazio.hide()
            self.table_trabalho()

    def atualizar_cpf_cliente(self):
        # Este método é chamado sempre que a combobox é ativada
        index = self.cmb_cliente.currentIndex()
        if index != -1:
            cpf_cliente = self.cmb_cliente.itemData(index)
            print(f"Cliente: {self.cmb_cliente.currentText()}, CPF: {cpf_cliente}")
        else:
            print("Nenhum cliente escolhido")

        # Conectar ao banco de dados
        db = Database()
        db.connect()

    def carregar_nome_na_combobox_cliente(self):
        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar funcionários do banco de dados
        clientes = db.get_cliente()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        self.cmb_cliente.clear()


        # Adicionar funcionários à combobox
        for cliente in clientes:
            cpf_cliente, nome_cliente = cliente  # Assuming funcionario is a tuple with (ID, Nome)
            # Add the name to the ComboBox and set its data to be the ID
            self.cmb_cliente.addItem(nome_cliente, userData=cpf_cliente)
            

    def atualizar_id_funcionario(self):
        # Este método é chamado sempre que a combobox é ativada
        index = self.cmb_funcionario_responsavel.currentIndex()
        if index != -1:
            id_funcionario = self.cmb_funcionario_responsavel.itemData(index)
            print(f"Funcionário: {self.cmb_funcionario_responsavel.currentText()}, ID: {id_funcionario}")
        else:
            print("No employee selected")

        # Conectar ao banco de dados
        db = Database()
        db.connect()

    def carregar_usuario_na_combobox_funcionario(self):
        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar funcionários do banco de dados
        funcionarios = db.get_funcionario()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        # Limpar a combobox antes de adicionar novos itens
        self.cmb_funcionario_responsavel.clear()

        # Adicionar funcionários à combobox
        for funcionario in funcionarios:
            id_funcionario, nome_funcionario = funcionario  # Assuming funcionario is a tuple with (ID, Nome)
            # Add the name to the ComboBox and set its data to be the ID
            self.cmb_funcionario_responsavel.addItem(nome_funcionario, userData=id_funcionario) 

    def setup_connections_trabalhos_filtro(self):
        self.txt_filtro_trabalhos.textChanged.connect(self.filter_table_trabalhos)

    def filter_table_trabalhos(self, text):
        for row in range(self.tw_trabalhos_geral.topLevelItemCount()):
            item = self.tw_trabalhos_geral.topLevelItem(row)
            nome_index_trabalho = 5 #Coluna onde está o nome do funcionário

            if item.text(nome_index_trabalho).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def table_trabalho(self):
        # Clear existing items in the QTreeWidget
        self.tw_trabalhos.clear()

        # Connect to the database and execute the query
        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                t.modelo_automovel, t.preco
            FROM trabalhos t
            JOIN funcionario f ON t.ID_funcionario = f.ID
            JOIN cliente c ON t.CPF_cliente = c.CPF
            WHERE t.dia_realizado IS NULL
        """
        result = pd.read_sql_query(query, cn)

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos, [str(row[col]) for col in result.columns])
            item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            # Cria e seta um botão de delete
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            # Obter id_trabalho para esta iteração
            id_trabalho = str(row['id_trabalho'])

            # Conectar o botão de delete à função confirmar_delete_trabalho
            delete_button.clicked.connect(functools.partial(self.excluir_trabalho, id_trabalho))

            # Configurar o botão de delete no item da árvore
            self.tw_trabalhos.setItemWidget(item, self.tw_trabalhos.columnCount() - 1, delete_button)

            # Cria o botão de editar na quarta coluna
            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            # Aqui está a lógica ajustada para o botão "Editar"
            edit_button.clicked.connect(lambda id_trabalho=id_trabalho: self.edit_selected_trabalho(id_trabalho))
            edit_button.clicked.connect(self.carregar_usuario_na_combobox_funcionario)
            edit_button.clicked.connect(self.carregar_nome_na_combobox_cliente)

            self.tw_trabalhos.setItemWidget(item, self.tw_trabalhos.columnCount() - 2, edit_button)

        self.tw_trabalhos.setSortingEnabled(False)

        for i in range(self.tw_trabalhos.columnCount()):
            self.tw_trabalhos.resizeColumnToContents(i)

    def table_trabalhos_geral(self):
        # Clear existing items in the QTreeWidget
        self.tw_trabalhos_geral.clear()

        # Connect to the database and execute the query
        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                t.modelo_automovel, t.preco, t.dia_realizado
            FROM trabalhos t
            JOIN funcionario f ON t.ID_funcionario = f.ID
            JOIN cliente c ON t.CPF_cliente = c.CPF
        """
        result = pd.read_sql_query(query, cn)

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos_geral, [str(row[col]) for col in result.columns])

        self.tw_trabalhos_geral.setSortingEnabled(False)

        for i in range(self.tw_trabalhos.columnCount()):
            self.tw_trabalhos_geral.resizeColumnToContents(i)

    def table_trabalhos_realizados(self):
        cn = sqlite3.connect('AutoCenter.db')
        query = """
            SELECT id_trabalho, nome_trabalho, id_funcionario, CPF_cliente,
                modelo_automovel, preco, dia_realizado
            FROM trabalhos WHERE dia_realizado IS NOT NULL
        """
        result = pd.read_sql_query(query, cn)

        # Clear the existing tree
        self.tw_trabalhos_realizados.clear()

        for index, row in result.iterrows():
            item = QTreeWidgetItem(self.tw_trabalhos_realizados, [str(row[col]) for col in result.columns])
            item.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

        self.tw_trabalhos_realizados.setSortingEnabled(False)

        # Adjust the resizing of columns
        for i in range(self.tw_trabalhos_realizados.columnCount()):
            self.tw_trabalhos_realizados.resizeColumnToContents(i)

        self.calcular_soma_precos()

    def excluir_trabalho(self, id_trabalho):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Trabalho", "Deseja realmente excluir o trabalho?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o cliente do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM trabalhos WHERE id_trabalho = ?", (id_trabalho,))
            connection.commit()
            connection.close()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Trabalho")
            msg.setText("Trabalho Excluído Com Sucesso")
            msg.exec_()
            # Atualizar a tabela de trabalhos após a exclusão
            self.table_trabalho()

    def edit_selected_trabalho(self, id_trabalho):
        selected_items = self.tw_trabalhos.selectedItems()

        # Verifica se mais de um item está selecionado
        if len(selected_items) > 1:
            QMessageBox.warning(self, "Alerta", "Selecione apenas um item para editar.")
            return

        # Perguntar ao usuário se eles desejam editar o cliente selecionado
        reply = QMessageBox.question(self, "Editar Trabalho", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o id do trabalho selecionado
            id_trabalho = selected_items[0].text(0)

            # Armazenar o id do trabalho em edição em um atributo da classe
            self.id_trabalho_edicao = id_trabalho

            # Obter os detalhes do trabalho no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            query = """
                SELECT t.id_trabalho, t.ID_funcionario, f.Nome AS nome_funcionario, 
                    t.CPF_cliente, c.Nome AS nome_cliente, t.nome_trabalho, 
                    t.modelo_automovel, t.preco
                FROM trabalhos t
                JOIN funcionario f ON t.ID_funcionario = f.ID
                JOIN cliente c ON t.CPF_cliente = c.CPF
                WHERE t.dia_realizado IS NULL AND t.id_trabalho = ?
            """
            cursor = connection.cursor()
            cursor.execute(query, (id_trabalho,))
            trabalho = cursor.fetchone()
            connection.close()

            self.updating_trabalho = True

            # Definir a página atual como a página de trabalhos
            self.Paginas.setCurrentWidget(self.page_AdicionarTrabalho)
            self.lbl_titulo_cadastro_trabalho.setText("Editar Trabalho")
            self.btn_salvar_trabalho.setText("Editar")

            # Definir o texto dos campos com os detalhes do cliente
            self.txt_nome_trabalho.setText(trabalho[5])

            # Configurar a combobox do funcionário responsável
            if self.cmb_funcionario_responsavel.count() > 0:
                # Utilizar o índice correto (2) para obter o nome do funcionário
                index = self.cmb_funcionario_responsavel.findText(trabalho[2])
                if index != -1:
                    self.cmb_funcionario_responsavel.setCurrentIndex(index)

            # Configurar a combobox do cliente
            if self.cmb_cliente.count() > 0:
                # Utilizar o índice correto (4) para obter o nome do cliente
                index = self.cmb_cliente.findText(trabalho[4])
                if index != -1:
                    self.cmb_cliente.setCurrentIndex(index)

            self.txt_modelo_auto.setText(trabalho[6])

            self.txt_preco_servico.setText(str(trabalho[7])) 
            self.txt_preco_servico.setText(trabalho[7])



    def update_trabalho(self, ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco):
        # Update the trabalho in the database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE trabalhos
            SET ID_funcionario=?, CPF_cliente=?, nome_trabalho=?, modelo_automovel=?, preco=?
            WHERE id_trabalho=?
        """, (ID_funcionario, CPF_cliente, nome_trabalho, modelo_automovel, preco, self.id_trabalho_edicao))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Cadastro de Trabalhos")
        msg.setText("Trabalho Editado com Sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_trabalho.clear()
        self.txt_modelo_auto.clear()
        self.txt_preco_servico.clear()
        self.table_trabalho()
        

    def realizar_trabalho(self):

        self.checked_items_out = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items_out.append(child.text(0))

        recurse(self.tw_trabalhos.invisibleRootItem())

        # Pergunta se o usuário realmente deseja fazer isso.
        self.question_trabalhos('realizar')

    def voltar_trabalho(self):

        self.checked_items = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items.append(child.text(0))

        recurse(self.tw_trabalhos_realizados.invisibleRootItem())
        self.question_trabalhos('voltar')   

    def calcular_soma_precos(self):
        # Inicialize a variável para armazenar a soma
        soma_precos_trabalho = 0.0

        # Iterar sobre os itens na árvore
        for item_index in range(self.tw_trabalhos_realizados.topLevelItemCount()):
            item = self.tw_trabalhos_realizados.topLevelItem(item_index)
            
            # A coluna correspondente ao preço depende da sua estrutura, ajuste conforme necessário
            # Aqui, assumimos que a coluna do preço é a sexta coluna (índice 5)
            preco_str = item.text(5)

            try:
                # Converter o valor do preço para float e somar à variável de soma
                preco = float(preco_str)
                soma_precos_trabalho += preco
            except ValueError:
                # Tratar o caso em que a conversão para float falha (pode ser necessário lidar com isso de acordo com seus requisitos)
                print(f"Erro ao converter preço: {preco_str}")

        self.txt_total_trabalhos.setText(str(soma_precos_trabalho))

    def question_trabalhos(self, table):

        msgBox = QMessageBox()

        if table == 'voltar':
            msgBox.setWindowTitle("Retornar Trabalho")
            msgBox.setText("Deseja retornar os trabalhos selecionados?")
            msgBox.setInformativeText("Os itens selecionados sairão para dos trabalhos feitos \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Trabalhos: {self.checked_items}")

        else:
            msgBox.setWindowTitle("Realizar Trabalho")
            msgBox.setText("Deseja realizar os trabalhos?")
            msgBox.setInformativeText("Os trabalhos sairão dos trabalhos disponíveis\n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Trabalhos: {self.checked_items_out}")

        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()

        if ret == QMessageBox.Yes:
            if table == "voltar":
                self.db = Database()
                print(f"Trabalhos selecionados para voltar: {self.checked_items}")
                self.db.connect()
                self.db.update_trabalho_disponivel(self.checked_items)
                self.db.close_connection()
                self.reset_table()
            else:
                dia_realizado = date.today()
                dia_realizado = dia_realizado.strftime('%d/%m/%Y')
                self.db = Database()
                self.db.connect()
                self.db.UpdateDate_trabalho(dia_realizado, self.checked_items_out)
                self.db.close_connection()
                self.reset_table()


    def CancelarTrabalho(self):
         self.txt_nome_trabalho.setText("")
         self.txt_modelo_auto.setText("")
         self.txt_ID.setText("")
         self.txt_preco_servico.setText("")


    def Adicionar_Cliente(self):
        if not self.txt_nome_cliente.text():
            self.txt_nome_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_cliente_vazio.show()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro do cliente")
            msg.exec_()
            return None
        else:
            self.txt_nome_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_cpf_cliente.text():
            self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cpf_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_cpf_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_telefone_cliente.text():
            self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_telefone_vazio.show()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_telefone_cliente.setStyleSheet(self.StyleLineEditOk)

        if not self.date_cliente.text():
            self.date_cliente.setStyleSheet(self.StyleLineEditErro)

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.date_cliente.setStyleSheet(self.StyleLineEditOk)

        # Se não houver erros, você consegue cadastrar ou editar o cliente
        CPF = self.txt_cpf_cliente.text()
        Nome = self.txt_nome_cliente.text()
        Telefone = self.txt_telefone_cliente.text()
        Dia_Agendamento = self.date_cliente.text()
        Peca_Requisitada = self.cmb_pecas.currentText()

        db = Database()
        db.connect()

        if not getattr(self, 'updating_cliente', False):
            # Verifica se o cliente já existe no banco de dados apenas se não estiver atualizando
            cliente_existe = db.check_if_cliente_exists(CPF)

            if cliente_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Cliente Existente")
                msg.setText(f"O cliente com CPF '{CPF}' já existe.")
                msg.exec_()
                self.txt_cpf_cliente.setText("")
                self.txt_nome_cliente.setText("")
                self.txt_telefone_cliente.setText("")
                self.lbl_erro_nome_cliente_vazio.hide()
                self.lbl_erro_cpf_vazio.hide()
                self.lbl_erro_telefone_vazio.hide()

        Chave_Nota = db.get_produto_id_by_descricao(Peca_Requisitada)

        if getattr(self, 'updating_cliente', False):
            # Marca a variável _updating_cliente como False após a atualização
            setattr(self, 'updating_cliente', False)
            # Chama a função de atualização
            self.update_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Nota, Dia_Agendamento)
        else:
            db.insert_cliente(CPF, Nome, Telefone, Dia_Agendamento, Peca_Requisitada, Chave_Nota)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Clientes")
            msg.setText("Cliente Adicionado com Sucesso")
            msg.exec_()
            self.txt_cpf_cliente.setText("")
            self.txt_nome_cliente.setText("")
            self.txt_telefone_cliente.setText("")
            self.lbl_erro_nome_cliente_vazio.hide()
            self.lbl_erro_cpf_vazio.hide()
            self.lbl_erro_telefone_vazio.hide()
            self.table_cliente()


    def atualizar_chave_nota(self):
        # Este método é chamado sempre que a combobox é ativada
        # Atualiza a variável Chave_Nota com base na descrição selecionada
        descricao_selecionada = self.cmb_pecas.currentText()

        db = Database()
        db.connect()
        self.Chave_Nota = db.get_produto_id_by_descricao(descricao_selecionada)
        db.close_connection()
        print(f"Descrição: {descricao_selecionada}, Chave_Nota: {self.Chave_Nota}")

    def carregar_produtos_na_combobox(self):

        # Conectar ao banco de dados
        db = Database()
        db.connect()

        # Recuperar produtos do banco de dados
        produtos = db.get_produtos()

        # Fechar a conexão com o banco de dados
        db.close_connection()

        # Limpar a combobox antes de adicionar novos itens
        self.cmb_pecas.clear()

        # Adicionar produtos à combobox
        for produto in produtos:
            descricao = produto[0]  # Substitua 'descricao' pelo nome correto da coluna no seu banco de dados
            self.cmb_pecas.addItem(descricao)
 
    def table_cliente(self):
        # Clear the existing items in the table
        self.tw_cliente.clear()

        # Create a database connection
        connection = sqlite3.connect("AutoCenter.db")
        # Create a cursor object
        cursor = connection.cursor()

        # Pega as informações da tabela cliente
        cursor.execute("SELECT CPF, Nome, Telefone, Dia_Agendamento, Peca_Requisitada, Chave_Notas FROM cliente ORDER BY CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento")
        records = cursor.fetchall()

        # Iterate over the records and add them to the tree widget
        for record in records:
            item = QTreeWidgetItem(self.tw_cliente)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            # Create and set the delete button for the sixth column
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_cliente(item, record))
            self.tw_cliente.setItemWidget(item, 6, delete_button)  # Ajuste na posição

            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                    "background-color: rgb(41, 49, 83); " \
                    "border: 2px solid rgb(55, 100, 183); " \
                    "border-radius: 4px; " \
                    "color: white; " \
                    "}" \
                    "QPushButton:pressed { " \
                    "border: 2px solid rgb(47, 82, 162); " \
                    "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.carregar_produtos_na_combobox())
            edit_button.clicked.connect(lambda: self.edit_selected_cliente())
            self.tw_cliente.setItemWidget(item, 7, edit_button)  # Ajuste na posição

        # Resize columns to their contents
        self.tw_cliente.resizeColumnToContents(0)  # Resize the first column for check states
        for i in range(1, self.tw_cliente.columnCount()):  # Resize remaining columns
            self.tw_cliente.resizeColumnToContents(i)
            self.tw_cliente.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_cliente.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        # Close the database connection
        connection.close()


    def setup_connections_cliente_filtro(self):
        self.txt_filtro_cliente.textChanged.connect(self.filter_table)

    def filter_table(self, text):
        for row in range(self.tw_cliente.topLevelItemCount()):
            item = self.tw_cliente.topLevelItem(row)
            nome_column_index = 1  # Substitua pelo índice da coluna "Nome"

            if item.text(nome_column_index).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def edit_selected_cliente(self):
        selected_items = self.tw_cliente.selectedItems()  # Obter itens selecionados na tabela
        
        # Verifica se mais de um item está selecionado
        if len(selected_items) > 1:
            # Exibe uma mensagem de alerta com ícone de alerta
            QMessageBox.warning(self, "Alerta", "Selecione apenas um item para editar.")
            return
        
        # Perguntar ao usuário se eles desejam editar o cliente selecionado
        reply = QMessageBox.question(self, "Editar Cliente", "Deseja editar o registro selecionado?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Obter o CPF do cliente selecionado
            cpf = selected_items[0].text(0) 

            # Armazenar o CPF do cliente em edição em um atributo da classe
            self.cpf_cliente_edicao = cpf

            # Obter os detalhes do cliente no banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM cliente WHERE CPF = ?", (cpf,))
            cliente = cursor.fetchone()
            connection.close()

            # Definir a página atual como a página de clientes
            self.Paginas.setCurrentWidget(self.page_clientes)

            # Definir o texto dos campos com os detalhes do cliente
            self.txt_nome_cliente.setText(cliente[1])
            self.txt_cpf_cliente.setText(cliente[0])
            self.txt_telefone_cliente.setText(cliente[2])

            if self.cmb_pecas.count() > 0:
                index = self.cmb_pecas.findText(cliente[4])
                if index != -1:
                    self.cmb_pecas.setCurrentIndex(index)

                date_string = cliente[3] 
                date = QDate.fromString(date_string, Qt.ISODate)
                self.date_cliente.setDate(date)
                self.updating_cliente = True
                self.lbl_titulo_cadastro_cliente.setText("Atualizar Cliente")
                self.btn_salvar_cliente.setText("Atualizar")


    def update_cliente(self, CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento):
        # Update the cliente in the database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE cliente
            SET CPF=?, Nome=?, Telefone=?, Peca_Requisitada=?, Chave_Notas=?, Dia_Agendamento=?
            WHERE CPF=?
        """, (CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento, CPF))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Cadastro de Cliente")
        msg.setText("Cliente Editado com Sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_cliente.clear()
        self.txt_telefone_cliente.clear()
        self.date_cliente.clear()
        self.table_cliente()

    def on_btn_salvar_cliente_clicked(self):
        # Get the values from the text boxes
        CPF = self.txt_cpf_cliente.text()
        Nome = self.txt_nome_cliente.text()
        Telefone = self.txt_telefone_cliente.text()
        Peca_Requisitada = self.cmb_pecas.currentText()
        Dia_Agendamento = self.date_cliente.text()

        # Check if a cliente is selected
        selected_items = self.tw_cliente.selectedItems()
        print("Selected Items:", selected_items)

        if selected_items:
            # Se o cliente estiver selecionado, edita
            CPF = selected_items[0].text(0)

            # Obter a Chave_Notas (anteriormente CPF_Cliente) da peça solicitada
            db = Database()
            db.connect()
            Chave_Notas = db.get_produto_id_by_descricao(Peca_Requisitada)
            db.close_connection()

            self.update_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento)
        else:
            # Se nenhum cliente estiver selecionado, salva normalmente
            self.insert_cliente(CPF, Nome, Telefone, Peca_Requisitada, Chave_Notas, Dia_Agendamento)

    def excluir_cliente(self, CPF):

        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Cliente", "Deseja realmente excluir o cliente?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o cliente do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM cliente WHERE CPF=?", (CPF,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de fornecedores após a exclusão
            self.table_cliente()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Cliente")
            msg.setText("Cliente Excluído Com Sucesso")
            msg.exec_()

    def confirmar_delete_cliente(self, item, record):

        # Obter o CPF do cliente a ser excluído
        CPF = record[0]

        # Chamar a função para excluir o cliente
        self.excluir_cliente(CPF)


    def eventFilter(self, obj, event):
        if obj is self.txt_cep_fornecedor and event.type() == QEvent.KeyPress and event.key() == Qt.Key_Return:
            self.CepFornecedor()
            return True
        return super().eventFilter(obj, event)

    def CepFornecedor(self):
        cep = self.txt_cep_fornecedor.text()
        resultado = brazilcep.get_address_from_cep(cep)

        if resultado:
            self.txt_logradouro_fornecedor.setText(resultado["street"])
            self.txt_bairro_fornecedor.setText(resultado["district"])
        
            cidade = resultado["city"]
            estado = resultado["uf"]
            self.txt_UF_fornecedor.setText(f"{cidade} / {estado}")
        else:
            print("CEP inválido")


    def Adicionar_fornecedor(self):

        if not self.txt_nome_fantasia.text():
            self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_nome_fantasia_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_nome_fantasia.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_razao_social.text():
            self.txt_razao_social.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_razao_social_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_razao_social.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_cnpj_fornecedor.text():
            self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cnpj_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_cnpj_fornecedor.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_IE_fornecedor.text():
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_IE_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditOk) 

        # Verificar se o preço é um valor numérico
        try:
            IE = float(self.txt_IE_fornecedor.text())
        except ValueError:
            # Se a conversão falhar, exibir mensagem de erro e retornar
            self.txt_IE_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.txt_IE_fornecedor.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("A Inscrição Estadual deve ser um valor numérico.")
            msg.exec_()
            return None
        else:
            # Se a conversão for bem-sucedida, continuar com o cadastro
            self.txt_preco_servico.setStyleSheet(self.StyleLineEditOk)
            self.lbl_erro_preco_vazio.hide()       
        if not self.txt_telefone_fornecedor.text():
            self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_telefone_fornecedor_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_telefone_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_email_fornecedor.text():
            self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_email_fornecedor_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_email_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_numero_fornecedor.text():
            self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_numero_casa_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_numero_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_logradouro_fornecedor.text():
            self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_logradouro_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_logradouro_fornecedor.setStyleSheet(self.StyleLineEditOk)     

        if not self.txt_bairro_fornecedor.text():
            self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_bairro_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_bairro_fornecedor.setStyleSheet(self.StyleLineEditOk)           

        if not self.txt_UF_fornecedor.text():
            self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_cidade_vazio.show()
   
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("Não foi possível fazer o cadastro")
            msg.exec_()
            return None
        else:
            self.txt_UF_fornecedor.setStyleSheet(self.StyleLineEditOk) 
       

        # Separa cidade e estado
        uf_fornecedor = self.txt_UF_fornecedor.text()
        if '/' in uf_fornecedor:
            cidade, estado = map(str.strip, uf_fornecedor.split('/'))
        else:
            cidade, estado = uf_fornecedor, ''  # Defina estado como vazio

        nome_fantasia = self.txt_nome_fantasia.text()
        razao_social = self.txt_razao_social.text()
        cnpj = self.txt_cnpj_fornecedor.text()
        insc_estadual = self.txt_IE_fornecedor.text()
        telefone = self.txt_telefone_fornecedor.text()
        email = self.txt_email_fornecedor.text()
        cep = self.txt_cep_fornecedor.text()
        numero = self.txt_numero_fornecedor.text()
        logradouro = self.txt_logradouro_fornecedor.text()
        bairro = self.txt_bairro_fornecedor.text()

        db = Database()
        db.connect()
    
        if not getattr(self, 'updating_fornecedor', False):
            # Verifica se o cliente já existe no banco de dados apenas se não estiver atualizando
            fornecedor_existe = db.check_if_fornecedor_exists(cnpj)

            if fornecedor_existe:
                db.close_connection()

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Fornecedor Existente")
                msg.setText(f"O fornecedor '{nome_fantasia}' já existe.")
                msg.exec_()
                nome_fantasia = self.txt_nome_fantasia.setText("")
                razao_social = self.txt_razao_social.setText("")
                cnpj = self.txt_cnpj_fornecedor.setText("")
                insc_estadual = self.txt_IE_fornecedor.setText("")
                telefone = self.txt_telefone_fornecedor.setText("")
                email = self.txt_email_fornecedor.setText("")
                cep = self.txt_cep_fornecedor.setText("")
                numero = self.txt_numero_fornecedor.setText("")
                logradouro = self.txt_logradouro_fornecedor.setText("")
                bairro = self.txt_bairro_fornecedor.setText("")

        # Marca a variável _updating_funcionario como False após a atualização
        if getattr(self, 'updating_fornecedor', False):
            setattr(self, 'updating_fornecedor', False)
            cnpj = self.tw_fornecedor.selectedItems()[0].text(0)
            self.update_fornecedor(cnpj, nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado)

        else:
                db.insert_fornecedor(nome_fantasia, razao_social,cnpj, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Cadastro de Fornecedor")
                msg.setText("Fornecedor Adicionado com Sucesso")
                msg.exec_()
                nome_fantasia = self.txt_nome_fantasia.setText("")
                razao_social = self.txt_razao_social.setText("")
                cnpj = self.txt_cnpj_fornecedor.setText("")
                insc_estadual = self.txt_IE_fornecedor.setText("")
                telefone = self.txt_telefone_fornecedor.setText("")
                email = self.txt_email_fornecedor.setText("")
                cep = self.txt_cep_fornecedor.setText("")
                numero = self.txt_numero_fornecedor.setText("")
                logradouro = self.txt_logradouro_fornecedor.setText("")
                bairro = self.txt_bairro_fornecedor.setText("")
                self.table_fornecedor()


    def setup_connections_fornecedor_filtro(self):
        self.txt_filtro_fornecedor.textChanged.connect(self.filter_table_fornecedor)

    def filter_table_fornecedor(self, text):
        for row in range(self.tw_fornecedor.topLevelItemCount()):
            item = self.tw_fornecedor.topLevelItem(row)
            nome_index_fornecedor = 1 #Coluna onde está o nome do funcionário

            if item.text(nome_index_fornecedor).lower().startswith(text.lower()):
                item.setHidden(False)
            else:
                item.setHidden(True)

    def table_fornecedor(self):
        # Clear the existing items in the table
        self.tw_fornecedor.clear()

        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()

        #Faz uma busca na tabela
        cursor.execute("SELECT cnpj, nome_fantasia, email, telefone FROM fornecedor ORDER BY cnpj, nome_fantasia, email, telefone")
        records = cursor.fetchall()

        # Interage com os records e grava eles na TreeWidget
        for record in records:
            item = QTreeWidgetItem(self.tw_fornecedor)
            for i in range(len(record)):
                item.setText(i, str(record[i]))

            # Cria e seta um botão de delete
            delete_button = QPushButton("Excluir")
            style = "QPushButton { " \
                                "background-color: rgb(41, 49, 83); " \
                                "border: 2px solid rgb(55, 100, 183); " \
                                "border-radius: 4px; " \
                                "color: white; " \
                                "}" \
                                "QPushButton:pressed { " \
                                "border: 2px solid rgb(47, 82, 162); " \
                                "}"
            delete_button.setStyleSheet(style)

            delete_button.clicked.connect(lambda item=item, record=record: self.confirmar_delete_fornecedor(item, record))
            self.tw_fornecedor.setItemWidget(item, self.tw_fornecedor.columnCount() - 1, delete_button)

            # Cria o botão de editar na quarta coluna
            edit_button = QPushButton("Editar")
            style = "QPushButton { " \
                                        "background-color: rgb(41, 49, 83); " \
                                        "border: 2px solid rgb(55, 100, 183); " \
                                        "border-radius: 4px; " \
                                        "color: white; " \
                                        "}" \
                                        "QPushButton:pressed { " \
                                        "border: 2px solid rgb(47, 82, 162); " \
                                        "}"
            edit_button.setStyleSheet(style)

            edit_button.clicked.connect(lambda: self.edit_selected_fornecedor())
            self.tw_fornecedor.setItemWidget(item, 4, edit_button)  # Ajuste para a nova posição da coluna

        # Resize columns to their contents
        self.tw_fornecedor.resizeColumnToContents(0)  # Resize the first column for check states
        for i in range(1, self.tw_cliente.columnCount()):  # Resize remaining columns
            self.tw_cliente.resizeColumnToContents(i)
            self.tw_cliente.header().resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self.tw_cliente.header().setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)

        # Close the database connection
        connection.close()

    def excluir_fornecedor(self, cnpj):
        # Confirmar exclusão com o usuário
        reply = QMessageBox.question(
            self, "Excluir Fornecedor", "Deseja realmente excluir o fornecedor?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Excluir o fornecedor do banco de dados
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("DELETE FROM fornecedor WHERE cnpj=?", (cnpj,))
            connection.commit()
            connection.close()

            # Atualizar a tabela de fornecedores após a exclusão
            self.table_fornecedor()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Exclusão de Fornecedor")
            msg.setText("Fornecedor excluído com sucesso")
            msg.exec_()

    def confirmar_delete_fornecedor(self, record):
        # Obter o ID do fornecedor a ser excluído
        cnpj = record[0]

        # Chamar a função para excluir o fornecedor
        self.excluir_fornecedor(cnpj)

    def edit_selected_fornecedor(self):
        selected_items = self.tw_fornecedor.selectedItems()  #Pega os itens selecionados na tabela
        if not selected_items:
            return  # Nenhum item selecionado

        # pergunta ao usuário se ele quer editar o fornecedor
        reply = QMessageBox.question(self, "Editar Fornecedor", "Deseja editar o registro selecionado?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # Pega o id do fornecedor selecionado
            cnpj = selected_items[0].text(0)  
            # Fetch os detalhes do fornecedor
            connection = sqlite3.connect("AutoCenter.db")
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM fornecedor WHERE cnpj = ?", (cnpj,))
            fornecedor = cursor.fetchone()
            connection.close()

            if fornecedor:
                #Manda o usuário para a tela de adicionar fornecedor
                self.Paginas.setCurrentWidget(self.page_adicionar_fornecedor)

                #Preenche as textboxes com os dados do fornecedor
                self.txt_nome_fantasia.setText(fornecedor[0]) 
                self.txt_razao_social.setText(fornecedor[1]) 
                self.txt_cnpj_fornecedor.setText(fornecedor[2])
                self.txt_IE_fornecedor.setText(str(fornecedor[3]))
                self.txt_telefone_fornecedor.setText(fornecedor[4])
                self.txt_email_fornecedor.setText(fornecedor[5])
                self.txt_cep_fornecedor.setText(fornecedor[6]) 
                self.txt_numero_fornecedor.setText(fornecedor[7])
                self.txt_logradouro_fornecedor.setText(fornecedor[8]) 
                self.txt_bairro_fornecedor.setText(fornecedor[9])
                self.txt_UF_fornecedor.setText(fornecedor[10] + " / " + fornecedor[11]) 

            self.updating_fornecedor = True
            self.lbl_titulo_fornecedor.setText("Editar Fornecedor")
            self.btn_salvar_fornecedor.setText("Editar")


    def update_fornecedor(self, cnpj, nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado):
        # Faz update do fornecedor no database
        connection = sqlite3.connect("AutoCenter.db")
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE fornecedor
            SET nome_fantasia=?, razao_social=?, insc_estadual=?, telefone=?, email=?, cep=?, numero=?, logradouro=?, bairro=?, cidade=?, estado=?
            WHERE cnpj=?
        """, (nome_fantasia, razao_social, insc_estadual, telefone, email, cep, numero, logradouro, bairro, cidade, estado, cnpj))
        connection.commit()
        connection.close()

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Edição de Fornecedor")
        msg.setText("Fornecedor editado com sucesso")
        msg.exec_()

        # Limpa os campos de entrada
        self.txt_nome_fantasia.clear()
        self.txt_razao_social.clear()
        self.txt_cnpj_fornecedor.clear()
        self.txt_IE_fornecedor.clear()
        self.txt_telefone_fornecedor.clear()
        self.txt_email_fornecedor.clear()
        self.txt_cep_fornecedor.clear()
        self.txt_numero_fornecedor.clear()
        self.table_fornecedor()


    def Adicionar_Item(self):
        if not self.txt_codigo_produto.text():
            self.txt_codigo_produto.setStyleSheet(self.StyleLineEditErro) 
            self.lbl_erro_codigo_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_codigo_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_quantidade_produto.text():
            self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_qtd_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_quantidade_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_descricao_produto.text():
            self.txt_descricao_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_descricao_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_descricao_produto.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_unidade_medida.text():
            self.txt_unidade_medida.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_unidade_medida_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_unidade_medida.setStyleSheet(self.StyleLineEditOk)

        if not self.txt_preco_produto.text():
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_valor_produto_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro ao Adicionar Produto")
            msg.setText("Não foi possível adicionar o produto")
            msg.exec_()
            return None
        else:
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk) 

        try:
            valorProd = float(self.txt_preco_produto.text())
        except ValueError:
            # Se a conversão falhar, exibir mensagem de erro e retornar
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditErro)
            self.lbl_erro_preco_vazio.show()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro no Cadastro")
            msg.setText("O preço do produto deve ser um valor numérico.")
            msg.exec_()
            return None   
        else:
            self.txt_preco_produto.setStyleSheet(self.StyleLineEditOk) 
            self.lbl_erro_preco_produto_vazio.hide()

        NFeRandom = random.randint(1, 100)
        NFe = str(NFeRandom) 
        serieRandom = random.randint(1, 100)
        serie = str(serieRandom)
        data_emissao = "//"
        chaveRandom = random.randint(1,100)
        chave = str(chaveRandom)
        cnpj_emitente = "850000"
        nome_emitente =  "Sigma Systems"
        valorNFe = "58,00"
        itemNotaRandom = random.randint(1, 100)
        itemNota = str(itemNotaRandom)
        cod = self.txt_codigo_produto.text()
        qntd = self.txt_quantidade_produto.text()
        descricao = self.txt_descricao_produto.text()
        unidade_medida = self.txt_unidade_medida.text()
        valorProd = self.txt_preco_produto.text()
        data_importacao = date.today()
        data_importacao = data_importacao.strftime('%d/%m/%Y')
        data_saida = ""

        db = Database()
        db.connect()

        produto_existe = db.check_if_produto_exists(cod)

        if produto_existe:
            db.close_connection()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Produto Existente")
            msg.setText(f"O Produto com código'{cod}' já existe.")
            msg.exec_()
            self.txt_codigo_produto.setText("")
            self.txt_quantidade_produto.setText("")
            self.txt_descricao_produto.setText("")
            self.txt_unidade_medida.setText("")
            self.txt_preco_produto.setText("")
            self.lbl_erro_codigo_produto_vazio.hide()
            self.lbl_erro_qtd_produto_vazio.hide()
            self.lbl_erro_descricao_produto_vazio.hide()
            self.lbl_erro_unidade_medida_vazio.hide()
            self.lbl_erro_preco_produto_vazio.hide()
        else:
            db.inserir_produtos(NFe, serie, data_emissao, chave, cnpj_emitente, nome_emitente,
                                valorNFe, itemNota, cod, qntd, descricao, unidade_medida, valorProd,
                                data_importacao, data_saida)
            db.close_connection()
            self.reset_table()

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Cadastro de Produtos")
            msg.setText("Produto adicionado com sucesso")
            msg.exec()

            self.txt_codigo_produto.setText("")
            self.txt_quantidade_produto.setText("")
            self.txt_descricao_produto.setText("")
            self.txt_unidade_medida.setText("")
            self.txt_preco_produto.setText("")
            self.lbl_erro_codigo_produto_vazio.hide()
            self.lbl_erro_qtd_produto_vazio.hide()
            self.lbl_erro_descricao_produto_vazio.hide()
            self.lbl_erro_unidade_medida_vazio.hide()
            self.lbl_erro_preco_produto_vazio.hide()


    def CancelarProduto(self):
         self.txt_codigo_produto.setText("")
         self.txt_quantidade_produto.setText("")
         self.txt_descricao_produto.setText("")
         self.txt_unidade_medida.setText("")
         self.txt_preco_produto.setText("")


    def open_path(self):
        self.path = QFileDialog.getExistingDirectory(self,str("Open Directory"),
                                                            "/home",
                                                            QFileDialog.ShowDirsOnly
                                                            | QFileDialog.DontResolveSymlinks)
        self.txt_arquivo_xml.setText(self.path)
    
    def import_arquivos_xml(self):

        xml = Read_xml(self.path)
        all = xml.all_files()

        db = Database()
        db.connect()

        if len(all) == 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Importação XML")
            msg.setText("Não foi possível concluir a importação. A pasta selecionada não possui um arquivo XML NF-e.")
            msg.exec_()
            return

        # try/except para capturar erros
        try:
            for i in all:
                fullDataSet = xml.nfe_data(i)
                db.insert_data(fullDataSet)
        except sqlite3.IntegrityError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Importação XML")
            msg.setText("Nota já existe no banco")
            msg.exec_()
        
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Importação XML")
        msg.setText("Importação Concluída")
        msg.exec_()
        self.reset_table()
        db.close_connection()
            

    def table_estoque(self):
        cn  =  sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("SELECT * FROM Notas WHERE data_saida = ''", cn)
        result = result.values.tolist()

        self.x = ""

        for i in result:
            #faz o check para identificar a mesma nota e adicionar um nivel
            if i[0] == self.x:
                QTreeWidgetItem(self.campo, i)
            else:
                self.campo = QTreeWidgetItem(self.tw_estoque, i)
                self.campo.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            self.x = i[0]

        self.tw_estoque.setSortingEnabled(False)
        self.tw_estoque.itemChanged.connect(self.botões)
   
        for i in range(1,14):
            self.tw_estoque.resizeColumnToContents(i)


    def botões(self):
        # Inicialize a lista de itens selecionados
        self.checked_items_out = []
    
        # Chame a função recursiva para verificar os estados de verificação
        self.recurse(self.tw_estoque.invisibleRootItem())
    
        # Verifique se a lista de itens selecionados não está vazia para mostrar os botões
        if self.checked_items_out:
            self.btn_excluir.show()
        else:
            self.btn_excluir.hide()

    def recurse(self, parent_item):
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            grand_children = child.childCount()
            if grand_children > 0:
                self.recurse(child)
            if child.checkState(0) == QtCore.Qt.Checked:
                self.checked_items_out.append(child.text(0))

    def questionExcluir(self):
        msgBox = QMessageBox()
        msgBox.setText("Deseja excluir os itens selecionados?")
        msgBox.setInformativeText("Os itens selecionados sairão do estoque para sempre.\nClique em 'Yes' para confirmar.")
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setDetailedText(f"Notas: {', '.join(self.checked_items_out)}")
        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()
    
        if ret == QMessageBox.Yes:
            self.db = Database()
            self.db.connect()
            self.db.excluir_itens(self.checked_items_out)
            self.db.close_connection()
            self.reset_table()
        else:
            self.db.close_connection()


    def table_saida(self):
        cn  =  sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("""SELECT NFe, serie, data_importacao, data_saida
         FROM Notas WHERE data_saida != ''""", cn)
        result = result.values.tolist()

        self.x = ""

        for i in result:
            #faz o check para identificar a mesma nota e adicionar um nivel
            if i[0] == self.x:
                QTreeWidgetItem(self.campo, i)
            else:
                self.campo = QTreeWidgetItem(self.tw_saida, i)
                self.campo.setCheckState(0, QtCore.Qt.CheckState.Unchecked)

            self.x = i[0]

        self.tw_saida.setSortingEnabled(False)
   
        for i in range(1,14):
            self.tw_saida.resizeColumnToContents(i)


    def table_geral(self):
        # Connect to the SQLite database
        db = QSqlDatabase("QSQLITE")
        db.setDatabaseName("AutoCenter.db")
        db.open()

        # Create a QSqlTableModel
        self.model = QSqlTableModel(db=db)

        # Set the table name
        self.model.setTable("Notas")

        # Select all data from the table
        self.model.select()

        # Set the model to the QTableView
        self.tb_geral.setModel(self.model)

        # Set the table headers
        self.tb_geral.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)


    def reset_table(self):
        self.tw_estoque.clear()
        self.tw_saida.clear()
        self.tw_cliente.clear()
        self.tw_fornecedor.clear()
        self.tw_trabalhos_realizados.clear()
        self.tw_trabalhos.clear()
        self.tw_trabalhos_geral.clear()

        self.table_saida()
        self.table_estoque()
        self.table_cliente()
        self.table_geral()
        self.table_trabalhos_geral()
        self.table_fornecedor()
        self.table_trabalho()
        self.table_trabalhos_realizados()


    def update_filter(self, s):
        s = re.sub("[\W_]+", "", s)
        filter_str = 'descricao LIKE "%{}%"'.format(s)
        self.model.setFilter(filter_str)


    def gerar_saida(self):

        self.checked_items_out = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items_out.append(child.text(0))
    
        recurse(self.tw_estoque.invisibleRootItem())

        #Pergunta se usuario realmente deseja fazer isso.
        self.question('saída')


    def gerar_estorno(self):
        
        self.checked_items = []

        def recurse(parent_item):
            for i in range(parent_item.childCount()):
                child = parent_item.child(i)
                grand_children = child.childCount()
                if grand_children > 0:
                    recurse(child)
                if child.checkState(0) == QtCore.Qt.Checked:
                    self.checked_items.append(child.text(0))

        recurse(self.tw_saida.invisibleRootItem())
        self.question('estorno')      


    def question(self, table):

        msgBox = QMessageBox()

        if table == 'estorno':
            msgBox.setText("Deseja estornar as notas selecionadas?")
            msgBox.setInformativeText("Os itens selecionados voltarão para o estoque \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Notas: {self.checked_items}")

        else:
            msgBox.setText("Deseja Gerar saída das nota selecionadas?")
            msgBox.setInformativeText("As notas abaixo sairão do estoque \n clique em 'Yes' para confirmar.")
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDetailedText(f"Notas: {self.checked_items_out}")

        msgBox.setIcon(QMessageBox.Question)
        ret = msgBox.exec_()

        if ret == QMessageBox.Yes:
            if table == "estorno":
                self.db = Database()
                self.db.connect()
                self.db.update_estorno(self.checked_items)
                self.db.close_connection()
                self.reset_table()
            else:
                data_saida = date.today()
                data_saida = data_saida.strftime('%d/%m/%Y')
                self.db = Database()
                self.db.connect()
                self.db.UpdateDate_estoque(data_saida, self.checked_items_out)
                self.db.close_connection()
                self.reset_table()


    def Converter_Excel(self):
        cnx = sqlite3.connect('AutoCenter.db')
        result = pd.read_sql_query("SELECT * FROM Notas", cnx)
        
        # Definir o caminho para a pasta raiz do projeto
        project_root = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(project_root, "Notas.xlsx")
        
        # Salvar o arquivo Excel
        result.to_excel(file_path, sheet_name='Notas', index=False)

        # Parte 2: Aplicar estilização ao arquivo Excel gerado
        self.estilizar_tabela_excel(file_path)

        # Mensagem de sucesso com ícone
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Notas em XML")
        msg.setText("Relatório gerado com sucesso!")
        
        # Adicionar ícone à QMessageBox
        icon_path = os.path.join(project_root, "Imagens", "excel.png")
        msg.setWindowIcon(QIcon(icon_path))


        msg.exec_()


    def estilizar_tabela_excel(self, file_path):
        # Abrir o arquivo Excel
        wb = load_workbook(file_path)
        ws = wb.active

        # Configurar um estilo personalizado para o cabeçalho
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Arial', color='FFFFFFFF', bold=True)  # Fonte Arial, cor branca para o texto
        header_style.fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')  # Cor ARGB para o preenchimento
        header_style.alignment = Alignment(horizontal='center', vertical='center')  # Centralizar texto

        thin_border = Border(left=Side(style='thin', color='FF6885bf'),
                            top=Side(style='thin', color='FF6885bf'),
                            right=Side(style='thin', color='FF6885bf'),
                            bottom=Side(style='thin', color='FF6885bf'))

        header_style.border = thin_border

        # Configurar um estilo personalizado para os campos
        cell_style = NamedStyle(name="cell_style")
        cell_style.font = Font(name='Arial', color='FFFFFFFF', bold=True)  # Fonte Arial, cor branca para o texto
        cell_style.fill = PatternFill(start_color='FF071325', end_color='FF071325', fill_type='solid')  # Cor ARGB para o preenchimento
        cell_style.border = thin_border
        cell_style.alignment = Alignment(horizontal='center', vertical='center')  # Centralizar texto

        # Aplicar o estilo ao cabeçalho da tabela
        for col_num, col in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1), start=1):
            for header_cell in col:
                header_cell.style = header_style

        # Alterar a formatação das células (exceto o cabeçalho)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            for cell in row:
                cell.style = cell_style

        # Autoajustar largura das colunas e altura das linhas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                try:
                    if len(str(cell.value)) > max_height:
                        max_height = len(cell.value)
                except:
                    pass
            adjusted_height = (max_height + 2)
            ws.row_dimensions[row[0].row].height = adjusted_height

        # Salvar as alterações
        wb.save(file_path)

    def onPrint(self):
        self.printDialog()


    def printDialog(self):

        filePath, filter = QFileDialog.getOpenFileName(self, 'Open file', '', 'PDF (*.pdf)')
        if not filePath:
            return
        file_extension = os.path.splitext(filePath)[1]

        if file_extension == ".pdf":
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                with tempfile.TemporaryDirectory() as path:
                    images = convert_from_path(filePath, dpi=300, output_folder=path, poppler_path=r'C:\Program Files\poppler-0.68.0\bin')
                    painter = QPainter()
                    painter.begin(printer)
                    for i, image in enumerate(images):
                        if i > 0:
                            printer.newPage()
                        rect = painter.viewport()
                        qtImage = ImageQt(image)
                        qtImageScaled = qtImage.scaled(rect.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        painter.drawImage(rect, qtImageScaled)
                    painter.end()
        else:
            pass

    def graphic(self):

        cnx = sqlite3.connect("AutoCenter.db")
        estoque = pd.read_sql_query('SELECT * FROM Notas', cnx)
        saida = pd.read_sql_query("SELECT * FROM Notas WHERE data_saida != ''", cnx)

        estoque = len(estoque)
        saida = len(saida)

        labels = "Estoque", "Saídas"
        sizes = [estoque, saida]
        fig1, axl = plt.subplots()
        axl.pie(sizes, labels=labels, autopct='%1.1555f%%', shadow=True, startangle=90)
        axl.axis("equal")

        plt.show()

    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Login()
    window.show()
    app.exec()